from __future__ import annotations

import os
import sys
import threading
import time
import webbrowser
from functools import wraps
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from flask import Flask, flash, g, redirect, render_template, request, send_file, session, url_for
from flask_migrate import Migrate
from dotenv import load_dotenv
from database import db, db_conn, IntegrityError, BASE_DIR as DATA_BASE_DIR, LOCAL_DB_PATH, database_url
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

load_dotenv()

APP_NAME = "Control de Inventario y Remisiones"


def resource_path(relative: str) -> Path:
    """Ruta compatible con ejecución normal y PyInstaller."""
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / relative


BASE_DIR = DATA_BASE_DIR
DB_PATH = LOCAL_DB_PATH
INVOICE_DIR = BASE_DIR / "invoices"
BACKUP_DIR = BASE_DIR / "backups"
INITIAL_XLSX = resource_path("data/Inventario_Inicial.xlsx")
SERVICES_XLSX = resource_path("data/Servicios_Iniciales.xlsx")

for folder in (INVOICE_DIR, BACKUP_DIR):
    folder.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "cambia-esta-clave-en-produccion")
app.config["SQLALCHEMY_DATABASE_URI"] = database_url()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}
db.init_app(app)
migrate = Migrate(app, db)


def money(value: Any) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0
    return "$ {:,.0f}".format(number).replace(",", ".")


app.jinja_env.filters["money"] = money


def parse_decimal(value: str | None, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    cleaned = str(value).strip().replace("$", "").replace(" ", "")
    if not cleaned:
        return default
    # Soporta 1.234.567,89 y 1234567.89
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Valor numérico inválido: {value}") from exc


def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            flash("Debes iniciar sesión.", "warning")
            return redirect(url_for("login", next=request.path))
        return view(**kwargs)
    return wrapped_view


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped_view(**kwargs):
            if g.user is None:
                flash("Debes iniciar sesión.", "warning")
                return redirect(url_for("login"))
            if g.user["role"] not in roles:
                flash("No tienes permisos para realizar esta acción.", "danger")
                return redirect(url_for("dashboard"))
            return view(**kwargs)
        return wrapped_view
    return decorator


def arturo_required(view):

    @wraps(view)
    def wrapped_view(**kwargs):

        if g.user is None:

            flash(
                "Debes iniciar sesión.",
                "warning"
            )

            return redirect(
                url_for("login")
            )


        full_name = str(
            g.user["full_name"] or ""
        ).strip().lower()

        role = str(
            g.user["role"] or ""
        ).strip().lower()


        if (
            full_name != "arturo lópez"
            or role != "administrador"
        ):

            flash(
                "Solo Arturo López puede editar remisiones.",
                "danger"
            )

            return redirect(
                url_for("dashboard")
            )


        return view(**kwargs)


    return wrapped_view


@app.before_request
def load_logged_in_user():
    user_id = session.get("user_id")
    if user_id is None:
        g.user = None
    else:
        with db_conn() as conn:
            g.user = conn.execute(
                "SELECT id, username, full_name, role, active FROM users WHERE id=?",
                (user_id,),
            ).fetchone()
        if g.user is None or not g.user["active"]:
            session.clear()
            g.user = None


def init_db() -> None:
    """Crea el esquema y carga datos predeterminados de forma aditiva."""
    with app.app_context():
        db.create_all()
        with db_conn() as conn:
            settings = conn.execute("SELECT * FROM settings WHERE id=1").fetchone()
            if not settings:
                conn.execute("""INSERT INTO settings(
                    id,business_name,nit,address,phone,email,invoice_prefix,next_invoice,tax_rate
                ) VALUES(1,?,?,?,?,?,?,?,?)""",
                ("Mi Negocio", "", "", "", "", "REM", 1, 0))
            else:
                conn.execute("UPDATE settings SET invoice_prefix='REM' WHERE id=1 AND invoice_prefix='FV'")

            default_users = [
                ("administrador", "admin123*", "Administrador", "administrador"),
                ("vendedor", "vendedor123*", "Vendedor genérico", "vendedor"),
                ("arturo.lopez", "Arturo1963*", "Arturo López", "administrador"),
            ]
            for username, password, full_name, role in default_users:
                if not conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
                    conn.execute(
                        "INSERT INTO users(username,password_hash,full_name,role,active) VALUES(?,?,?,?,1)",
                        (username, generate_password_hash(password), full_name, role),
                    )

            for seller_name in ("James", "Leonardo"):
                if not conn.execute("SELECT id FROM salespeople WHERE name=?", (seller_name,)).fetchone():
                    conn.execute("INSERT INTO salespeople(name,active) VALUES(?,1)", (seller_name,))

        seed_services_from_excel()
        seed_from_excel_if_empty()


def seed_services_from_excel() -> None:
    if not SERVICES_XLSX.exists():
        return
    wb = load_workbook(SERVICES_XLSX, data_only=True)
    added = 0
    with db_conn() as conn:
        for ws in wb.worksheets:
            # Encabezados útiles empiezan en la fila 3 en el archivo entregado.
            for row in ws.iter_rows(min_row=3, values_only=True):
                category = str(row[0] or "").strip()
                model = str(row[1] or "").strip()
                price = row[2] if len(row) > 2 else None
                if not category or not model or not isinstance(price, (int, float)):
                    continue
                name = f"{category.strip()} - {model.strip()}"
                if not conn.execute("SELECT id FROM services WHERE LOWER(name)=LOWER(?)", (name,)).fetchone():
                    conn.execute(
                        """INSERT INTO services(name,default_price,worker_percentage,business_percentage,active)
                           VALUES(?,?,70,30,1)""",
                        (name, float(price)),
                    )
                    added += 1
    if added:
        print(f"Servicios iniciales agregados: {added}")


def seed_from_excel_if_empty() -> None:
    with db_conn() as conn:
        row = conn.execute("SELECT COUNT(*) AS total FROM products").fetchone()
        if (row and int(row["total"]) > 0) or not INITIAL_XLSX.exists():
            return

        wb = load_workbook(INITIAL_XLSX, data_only=False)
        ws = wb["Inventario"]
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            code, description, category, unit, stock_initial = row[0], row[1], row[2], row[3], row[4]
            total_initial, min_stock, notes = row[8], row[11], row[13]
            if not code or not description:
                continue
            stock = float(stock_initial or 0)
            total = float(total_initial or 0)
            cost = total / stock if stock else 0
            code = str(code)
            if conn.execute("SELECT id FROM products WHERE code=?", (code,)).fetchone():
                continue
            conn.execute(
                """INSERT INTO products
                (code, description, category, unit, stock, min_stock, cost, sale_price, notes, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
                (code, str(description), str(category or ""), str(unit or "Unidad"), stock,
                 float(min_stock or 1), cost, cost, str(notes or "")),
            )
            imported += 1
        print(f"Productos importados: {imported}")


def get_settings(conn=None, lock: bool = False):
    sql = "SELECT * FROM settings WHERE id=1"
    if lock and db.engine.dialect.name == "postgresql":
        sql += " FOR UPDATE"
    if conn is not None:
        return conn.execute(sql).fetchone()
    with db_conn() as local:
        return local.execute(sql).fetchone()




@app.route("/healthz")
def healthz():
    return {"status": "ok", "database": db.engine.dialect.name}, 200


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.user is not None:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        with db_conn() as conn:
            user = conn.execute(
                "SELECT * FROM users WHERE username=? AND active=1", (username,)
            ).fetchone()
        if user is None or not check_password_hash(user["password_hash"], password):
            flash("Usuario o contraseña incorrectos.", "danger")
        else:
            session.clear()
            session["user_id"] = user["id"]
            return redirect(request.args.get("next") or url_for("dashboard"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/usuarios")
@role_required("administrador")
def users():
    with db_conn() as conn:
        rows = conn.execute(
            "SELECT id,username,full_name,role,active,created_at FROM users ORDER BY full_name"
        ).fetchall()
    return render_template("users.html", users=rows)


@app.route("/usuarios/nuevo", methods=["GET", "POST"])
@role_required("administrador")
def user_new():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        full_name = request.form.get("full_name", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role", "vendedor")
        try:
            if not username or not full_name:
                raise ValueError("El usuario y el nombre son obligatorios.")
            if len(password) < 8:
                raise ValueError("La contraseña debe tener al menos 8 caracteres.")
            if role not in ("administrador", "vendedor"):
                raise ValueError("Rol inválido.")
            with db_conn() as conn:
                conn.execute(
                    "INSERT INTO users(username,password_hash,full_name,role) VALUES(?,?,?,?)",
                    (username, generate_password_hash(password), full_name, role),
                )
            flash("Usuario creado correctamente.", "success")
            return redirect(url_for("users"))
        except IntegrityError:
            flash("Ese nombre de usuario ya existe.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    return render_template("user_form.html", user=None)


@app.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
@role_required("administrador")
def user_edit(user_id: int):
    with db_conn() as conn:
        user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("users"))

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "vendedor")
        active = 1 if request.form.get("active") == "1" else 0
        password = request.form.get("password", "")
        if role not in ("administrador", "vendedor"):
            flash("Rol inválido.", "danger")
            return render_template("user_form.html", user=user)
        if user_id == g.user["id"] and not active:
            flash("No puedes desactivar tu propio usuario.", "danger")
            return redirect(url_for("users"))
        with db_conn() as conn:
            if password:
                if len(password) < 8:
                    flash("La contraseña debe tener al menos 8 caracteres.", "danger")
                    return render_template("user_form.html", user=user)
                conn.execute(
                    "UPDATE users SET full_name=?,role=?,active=?,password_hash=? WHERE id=?",
                    (full_name, role, active, generate_password_hash(password), user_id),
                )
            else:
                conn.execute(
                    "UPDATE users SET full_name=?,role=?,active=? WHERE id=?",
                    (full_name, role, active, user_id),
                )
        flash("Usuario actualizado.", "success")
        return redirect(url_for("users"))
    return render_template("user_form.html", user=user)


@app.route("/")
@login_required
def dashboard():
    with db_conn() as conn:
        stats = conn.execute(
            """
            SELECT COUNT(*) product_count,
                   COALESCE(SUM(stock),0) total_units,
                   COALESCE(SUM(stock * cost),0) inventory_cost,
                   COALESCE(SUM(stock * sale_price),0) inventory_sale,
                   SUM(CASE WHEN stock <= 0 THEN 1 ELSE 0 END) out_count,
                   SUM(CASE WHEN stock > 0 AND stock <= min_stock THEN 1 ELSE 0 END) low_count
            FROM products WHERE active=1
            """
        ).fetchone()
        recent_invoices = conn.execute("SELECT * FROM invoices ORDER BY id DESC LIMIT 8").fetchall()
        low_products = conn.execute(
            "SELECT * FROM products WHERE active=1 AND stock <= min_stock ORDER BY stock ASC LIMIT 10"
        ).fetchall()
    return render_template("dashboard.html", stats=stats, recent_invoices=recent_invoices, low_products=low_products)


@app.route("/productos")
@login_required
def products():
    q = request.args.get("q", "").strip()
    with db_conn() as conn:
        if q:
            rows = conn.execute(
                """SELECT * FROM products WHERE active=1 AND
                (code LIKE ? OR description LIKE ? OR category LIKE ?)
                ORDER BY description""",
                (f"%{q}%", f"%{q}%", f"%{q}%"),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM products WHERE active=1 ORDER BY description").fetchall()
    return render_template("products.html", products=rows, q=q)


@app.route("/productos/nuevo", methods=["GET", "POST"])
@role_required("administrador")
def product_new():
    if request.method == "POST":
        try:
            values = (
                request.form["code"].strip(), request.form["description"].strip(),
                request.form.get("category", "").strip(), request.form.get("unit", "Unidad").strip(),
                float(parse_decimal(request.form.get("stock"))), float(parse_decimal(request.form.get("min_stock"), Decimal("1"))),
                float(parse_decimal(request.form.get("cost"))), float(parse_decimal(request.form.get("sale_price"))),
                request.form.get("notes", "").strip(),
            )
            with db_conn() as conn:
                conn.execute(
                    """INSERT INTO products(code, description, category, unit, stock, min_stock, cost, sale_price, notes)
                    VALUES(?,?,?,?,?,?,?,?,?)""", values
                )
            flash("Producto creado correctamente.", "success")
            return redirect(url_for("products"))
        except IntegrityError:
            flash("El código del producto ya existe.", "danger")
        except (ValueError, KeyError) as exc:
            flash(str(exc), "danger")
    return render_template("product_form.html", product=None)


@app.route("/productos/<int:product_id>/editar", methods=["GET", "POST"])
@role_required("administrador")
def product_edit(product_id: int):
    with db_conn() as conn:
        product = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    if not product:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("products"))

    if request.method == "POST":
        try:
            values = (
                request.form["code"].strip(), request.form["description"].strip(),
                request.form.get("category", "").strip(), request.form.get("unit", "Unidad").strip(),
                float(parse_decimal(request.form.get("min_stock"), Decimal("1"))),
                float(parse_decimal(request.form.get("cost"))), float(parse_decimal(request.form.get("sale_price"))),
                request.form.get("notes", "").strip(), product_id,
            )
            with db_conn() as conn:
                conn.execute(
                    """UPDATE products SET code=?, description=?, category=?, unit=?, min_stock=?, cost=?,
                    sale_price=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""", values
                )
            flash("Producto actualizado.", "success")
            return redirect(url_for("products"))
        except IntegrityError:
            flash("Ese código ya está asignado a otro producto.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    return render_template("product_form.html", product=product)

@app.route(
    "/productos/<int:product_id>/eliminar",
    methods=["POST"]
)
@role_required("administrador")
def product_delete(product_id: int):

    try:

        with db_conn() as conn:

            product = conn.execute(
                """
                SELECT *
                FROM products
                WHERE id=?
                """,
                (product_id,)
            ).fetchone()

            if not product:
                raise ValueError(
                    "Producto no encontrado."
                )


            invoice_usage = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM invoice_items
                WHERE product_id=?
                """,
                (product_id,)
            ).fetchone()


            movement_usage = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM movements
                WHERE product_id=?
                """,
                (product_id,)
            ).fetchone()


            has_history = (
                int(invoice_usage["total"] or 0) > 0
                or
                int(movement_usage["total"] or 0) > 0
            )


            # ============================================
            # PRODUCTO SIN HISTORIAL
            # Se puede borrar completamente
            # ============================================

            if not has_history:

                conn.execute(
                    """
                    DELETE FROM products
                    WHERE id=?
                    """,
                    (product_id,)
                )

                flash(
                    f"Producto '{product['description']}' "
                    f"eliminado completamente.",
                    "success"
                )


            # ============================================
            # PRODUCTO CON HISTORIAL
            # No borramos para no dañar remisiones
            # ============================================

            else:

                conn.execute(
                    """
                    UPDATE products

                    SET
                        active=0,
                        updated_at=CURRENT_TIMESTAMP

                    WHERE id=?
                    """,
                    (product_id,)
                )

                flash(
                    f"Producto '{product['description']}' "
                    f"retirado del inventario. "
                    f"Se conservó su historial.",
                    "success"
                )


        return redirect(
            url_for("products")
        )


    except ValueError as exc:

        flash(
            str(exc),
            "danger"
        )

        return redirect(
            url_for("products")
        )    


@app.route("/movimientos", methods=["GET", "POST"])
@role_required("administrador")
def movements():
    if request.method == "POST":
        try:
            product_id = int(request.form["product_id"])
            movement_type = request.form["movement_type"].strip()
            quantity = parse_decimal(request.form["quantity"])

            if quantity <= Decimal("0"):
                raise ValueError("La cantidad debe ser mayor que cero.")

            if movement_type not in ("Entrada", "Salida"):
                raise ValueError("Tipo de movimiento inválido.")

            signed = quantity if movement_type == "Entrada" else -quantity

            with db_conn() as conn:
                product_sql = "SELECT * FROM products WHERE id=?"
                if db.engine.dialect.name == "postgresql":
                    product_sql += " FOR UPDATE"

                product = conn.execute(product_sql, (product_id,)).fetchone()
                if not product:
                    raise ValueError("Producto no encontrado.")

                current_stock = Decimal(str(product["stock"] or 0))
                new_stock = current_stock + signed

                if new_stock < Decimal("0"):
                    raise ValueError("La salida supera el inventario disponible.")

                conn.execute(
                    """UPDATE products
                       SET stock=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (float(new_stock), product_id),
                )
                conn.execute(
                    """INSERT INTO movements(
                           product_id,movement_type,quantity,reference,notes,created_at,user_id
                       ) VALUES(?,?,?,?,?,?,?)""",
                    (
                        product_id,
                        movement_type,
                        float(quantity),
                        request.form.get("reference", "").strip(),
                        request.form.get("notes", "").strip(),
                        datetime.now().isoformat(timespec="seconds"),
                        g.user["id"],
                    ),
                )

            flash(
                f"Movimiento registrado correctamente. Nuevo inventario: {new_stock:g}",
                "success",
            )
            return redirect(url_for("movements"))

        except (ValueError, KeyError, TypeError) as exc:
            flash(str(exc), "danger")

    with db_conn() as conn:
        product_rows = conn.execute(
            """SELECT id,code,description,stock
               FROM products
               WHERE active=1
               ORDER BY description"""
        ).fetchall()
        rows = conn.execute(
            """SELECT m.*, p.code, p.description, u.full_name AS user_name
               FROM movements m
               JOIN products p ON p.id=m.product_id
               LEFT JOIN users u ON u.id=m.user_id
               ORDER BY m.id DESC
               LIMIT 100"""
        ).fetchall()

    return render_template(
        "movements.html",
        products=product_rows,
        movements=rows,
    )



@app.route("/vendedores")
@login_required
def salespeople():
    with db_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM salespeople ORDER BY active DESC, name"
        ).fetchall()
    return render_template("salespeople.html", salespeople=rows)


@app.route("/vendedores/nuevo", methods=["GET", "POST"])
@role_required("administrador")
def salesperson_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        try:
            if not name:
                raise ValueError("El nombre del vendedor es obligatorio.")
            with db_conn() as conn:
                conn.execute("INSERT INTO salespeople(name,active) VALUES(?,1)", (name,))
            flash("Vendedor agregado correctamente.", "success")
            return redirect(url_for("salespeople"))
        except IntegrityError:
            flash("Ya existe un vendedor con ese nombre.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    return render_template("salesperson_form.html", salesperson=None)


@app.route("/vendedores/<int:salesperson_id>/editar", methods=["GET", "POST"])
@role_required("administrador")
def salesperson_edit(salesperson_id: int):
    with db_conn() as conn:
        seller = conn.execute("SELECT * FROM salespeople WHERE id=?", (salesperson_id,)).fetchone()
    if not seller:
        flash("Vendedor no encontrado.", "danger")
        return redirect(url_for("salespeople"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        active = 1 if request.form.get("active") == "1" else 0
        try:
            if not name:
                raise ValueError("El nombre del vendedor es obligatorio.")
            with db_conn() as conn:
                conn.execute(
                    "UPDATE salespeople SET name=?,active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (name, active, salesperson_id),
                )
            flash("Vendedor actualizado.", "success")
            return redirect(url_for("salespeople"))
        except IntegrityError:
            flash("Ya existe otro vendedor con ese nombre.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    return render_template("salesperson_form.html", salesperson=seller)


@app.route("/servicios", methods=["GET", "POST"])
@login_required
def services():

    # =====================================================
    # CREAR SERVICIO
    # =====================================================

    if request.method == "POST":

        if g.user["role"] != "administrador":

            flash(
                "Solo el administrador puede crear servicios.",
                "danger"
            )

            return redirect(
                url_for("services")
            )

        try:

            name = request.form.get(
                "name",
                ""
            ).strip()

            price = float(
                parse_decimal(
                    request.form.get(
                        "default_price"
                    )
                )
            )

            if not name:

                raise ValueError(
                    "El nombre del servicio es obligatorio."
                )

            if price < 0:

                raise ValueError(
                    "El valor del servicio no puede ser negativo."
                )

            with db_conn() as conn:

                conn.execute(
                    """
                    INSERT INTO services(
                        name,
                        default_price,
                        worker_percentage,
                        business_percentage
                    )
                    VALUES(?,?,70,30)
                    """,
                    (
                        name,
                        price
                    )
                )

            flash(
                "Servicio creado correctamente.",
                "success"
            )

            return redirect(
                url_for("services")
            )

        except IntegrityError:

            flash(
                "Ya existe un servicio con ese nombre.",
                "danger"
            )

        except ValueError as exc:

            flash(
                str(exc),
                "danger"
            )


    # =====================================================
    # FILTROS
    # =====================================================

    date_from = request.args.get(
        "date_from",
        ""
    ).strip()

    date_to = request.args.get(
        "date_to",
        ""
    ).strip()

    worker_filter = request.args.get(
        "worker",
        ""
    ).strip()


    filters = []
    params = []


    if date_from:

        filters.append(
            "DATE(i.issued_at) >= ?"
        )

        params.append(
            date_from
        )


    if date_to:

        filters.append(
            "DATE(i.issued_at) <= ?"
        )

        params.append(
            date_to
        )


    if worker_filter:

        filters.append(
            "LOWER(TRIM(isi.worker_name)) = LOWER(?)"
        )

        params.append(
            worker_filter
        )


    where_clause = ""

    if filters:

        where_clause = (
            " WHERE "
            + " AND ".join(filters)
        )


    # =====================================================
    # CONSULTAS
    # =====================================================

    with db_conn() as conn:

        # ================================================
        # CATÁLOGO DE SERVICIOS
        # ================================================

        catalog = conn.execute(
            """
            SELECT *
            FROM services
            WHERE active=1
            ORDER BY name
            """
        ).fetchall()


        # ================================================
        # HISTORIAL FILTRADO
        # ================================================

        history = conn.execute(
            f"""
            SELECT
                isi.*,
                i.number,
                i.issued_at,
                i.customer_name,
                i.vehicle_plate

            FROM invoice_service_items isi

            JOIN invoices i
                ON i.id = isi.invoice_id

            {where_clause}

            ORDER BY
                i.issued_at DESC,
                isi.id DESC

            LIMIT 1000
            """,
            tuple(params)
        ).fetchall()


        # ================================================
        # RESUMEN GENERAL FILTRADO
        # ================================================

        summary = conn.execute(
            f"""
            SELECT

                COUNT(*) AS service_lines,

                COALESCE(
                    SUM(isi.line_total),
                    0
                ) AS service_total,

                COALESCE(
                    SUM(isi.worker_amount),
                    0
                ) AS worker_total,

                COALESCE(
                    SUM(isi.business_amount),
                    0
                ) AS business_total

            FROM invoice_service_items isi

            JOIN invoices i
                ON i.id = isi.invoice_id

            {where_clause}
            """,
            tuple(params)
        ).fetchone()


        # ================================================
        # DETALLE POR TRABAJADOR Y SERVICIO
        # ================================================

        worker_filters = list(filters)
        worker_params = list(params)

        worker_filters.append(
            "TRIM(isi.worker_name) <> ''"
        )

        worker_where = (
            " WHERE "
            + " AND ".join(
                worker_filters
            )
        )


        workers = conn.execute(
            f"""
            SELECT

                isi.worker_name,

                isi.service_name,

                COUNT(*) AS service_lines,

                COALESCE(
                    SUM(isi.quantity),
                    0
                ) AS quantity_total,

                COALESCE(
                    SUM(isi.line_total),
                    0
                ) AS service_total,

                COALESCE(
                    SUM(isi.worker_amount),
                    0
                ) AS worker_total,

                COALESCE(
                    SUM(isi.business_amount),
                    0
                ) AS business_total

            FROM invoice_service_items isi

            JOIN invoices i
                ON i.id = isi.invoice_id

            {worker_where}

            GROUP BY
                isi.worker_name,
                isi.service_name

            ORDER BY
                isi.worker_name ASC,
                isi.service_name ASC
            """,
            tuple(worker_params)
        ).fetchall()


        # ================================================
        # LISTADO DE TRABAJADORES PARA FILTRO
        # ================================================

        worker_options = conn.execute(
            """
            SELECT DISTINCT
                worker_name

            FROM invoice_service_items

            WHERE
                TRIM(worker_name) <> ''

            ORDER BY
                worker_name
            """
        ).fetchall()


        # ================================================
        # RESUMEN POR DÍA
        # ================================================

        daily = conn.execute(
            f"""
            SELECT

                DATE(i.issued_at) AS service_date,

                COUNT(*) AS service_lines,

                COALESCE(
                    SUM(isi.line_total),
                    0
                ) AS service_total,

                COALESCE(
                    SUM(isi.worker_amount),
                    0
                ) AS worker_total,

                COALESCE(
                    SUM(isi.business_amount),
                    0
                ) AS business_total

            FROM invoice_service_items isi

            JOIN invoices i
                ON i.id = isi.invoice_id

            {where_clause}

            GROUP BY
                DATE(i.issued_at)

            ORDER BY
                DATE(i.issued_at) DESC
            """,
            tuple(params)
        ).fetchall()


    return render_template(
        "services.html",

        services=catalog,

        history=history,

        summary=summary,

        workers=workers,

        worker_options=worker_options,

        daily=daily,

        date_from=date_from,

        date_to=date_to,

        worker_filter=worker_filter
    )

@app.route("/servicios/<int:service_id>/editar", methods=["GET", "POST"])
@role_required("administrador")
def service_edit(service_id: int):
    with db_conn() as conn:
        service = conn.execute("SELECT * FROM services WHERE id=?", (service_id,)).fetchone()
    if not service:
        flash("Servicio no encontrado.", "danger")
        return redirect(url_for("services"))

    if request.method == "POST":
        try:
            name = request.form.get("name", "").strip()
            price = float(parse_decimal(request.form.get("default_price")))
            active = 1 if request.form.get("active") == "1" else 0
            if not name:
                raise ValueError("El nombre del servicio es obligatorio.")
            if price < 0:
                raise ValueError("El valor del servicio no puede ser negativo.")
            with db_conn() as conn:
                conn.execute(
                    """UPDATE services SET name=?,default_price=?,active=?,updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (name, price, active, service_id),
                )
            flash("Servicio actualizado.", "success")
            return redirect(url_for("services"))
        except IntegrityError:
            flash("Ya existe otro servicio con ese nombre.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    return render_template("service_form.html", service=service)


@app.route("/facturas")
@login_required
def invoices():
    with db_conn() as conn:
        rows = conn.execute("""SELECT i.*, u.full_name AS user_name FROM invoices i
            LEFT JOIN users u ON u.id=i.user_id ORDER BY i.id DESC""").fetchall()
    return render_template("invoices.html", invoices=rows)


@app.route("/facturas/nueva", methods=["GET", "POST"])
@login_required
def invoice_new():

    form_data = request.form if request.method == "POST" else {}
    if request.method == "POST":
        try:
            customer_name = request.form.get("customer_name", "Consumidor final").strip() or "Consumidor final"
            customer_document = request.form.get("customer_document", "").strip()
            customer_phone = request.form.get("customer_phone", "").strip()
            customer_address = request.form.get("customer_address", "").strip()
            vehicle_plate = request.form.get("vehicle_plate", "").strip().upper()
            salesperson_id_raw = request.form.get("salesperson_id", "").strip()
            if not salesperson_id_raw:
                raise ValueError("Selecciona quién está realizando la venta.")
            salesperson_id = int(salesperson_id_raw)
            payment_method = request.form.get("payment_method", "Efectivo").strip()
            notes = request.form.get("notes", "").strip()

            credit_due_date_raw = request.form.get("credit_due_date", "").strip()
            credit_initial_payment = parse_decimal(
                request.form.get("credit_initial_payment", "0")
            )
            credit_initial_payment_method = request.form.get(
                "credit_initial_payment_method", "Efectivo"
            ).strip() or "Efectivo"

            if credit_initial_payment < 0:
                raise ValueError("El abono inicial no puede ser negativo.")

            credit_due_date = None
            if payment_method == "Crédito" and credit_due_date_raw:
                try:
                    credit_due_date = datetime.strptime(
                        credit_due_date_raw, "%Y-%m-%d"
                    ).date()
                except ValueError as exc:
                    raise ValueError("La fecha acordada de pago no es válida.") from exc

            product_ids = request.form.getlist("product_id[]")
            quantities = request.form.getlist("quantity[]")
            prices = request.form.getlist("price[]")

            service_ids = request.form.getlist("service_id[]")
            service_workers = request.form.getlist("service_worker[]")
            service_quantities = request.form.getlist("service_quantity[]")
            service_prices = request.form.getlist("service_price[]")

            # Filtrar filas vacías que puedan quedar en el formulario.
            product_rows_raw = [x for x in zip(product_ids, quantities, prices) if str(x[0]).strip()]
            service_rows_raw = [x for x in zip(service_ids, service_workers, service_quantities, service_prices) if str(x[0]).strip()]
            if not product_rows_raw and not service_rows_raw:
                raise ValueError("Agrega al menos un producto o un servicio a la remisión.")

            with db_conn() as conn:
                settings = get_settings(conn, lock=True)
                salesperson = conn.execute(
                    "SELECT id,name FROM salespeople WHERE id=? AND active=1",
                    (salesperson_id,),
                ).fetchone()
                if not salesperson:
                    raise ValueError("El vendedor seleccionado no está disponible.")
                number = f"{settings['invoice_prefix']}-{settings['next_invoice']:06d}"
                product_lines = []
                service_lines = []
                subtotal = Decimal("0")

                for product_id_raw, qty_raw, price_raw in product_rows_raw:
                    product_id = int(product_id_raw)
                    qty = parse_decimal(qty_raw)
                    price = parse_decimal(price_raw)
                    if qty <= 0 or price < 0:
                        raise ValueError("Revisa cantidades y precios de los productos.")
                    product_sql = "SELECT * FROM products WHERE id=? AND active=1"
                    if db.engine.dialect.name == "postgresql":
                        product_sql += " FOR UPDATE"
                    product = conn.execute(product_sql, (product_id,)).fetchone()
                    if not product:
                        raise ValueError("Uno de los productos ya no está disponible.")
                    if Decimal(str(product["stock"])) < qty:
                        raise ValueError(f"Stock insuficiente para {product['description']}. Disponible: {product['stock']}")
                    line_total = qty * price
                    subtotal += line_total
                    product_lines.append((product, qty, price, line_total))

                for service_id_raw, worker_name, qty_raw, price_raw in service_rows_raw:
                    service_id = int(service_id_raw)
                    worker_name = (worker_name or "").strip()
                    qty = parse_decimal(qty_raw)
                    price = parse_decimal(price_raw)
                    if not worker_name:
                        raise ValueError("Indica el trabajador responsable de cada servicio.")
                    if qty <= 0 or price < 0:
                        raise ValueError("Revisa cantidades y valores de los servicios.")
                    service = conn.execute("SELECT * FROM services WHERE id=? AND active=1", (service_id,)).fetchone()
                    if not service:
                        raise ValueError("Uno de los servicios ya no está disponible.")
                    line_total = qty * price
                    worker_pct = Decimal(str(service["worker_percentage"] or 70))
                    business_pct = Decimal(str(service["business_percentage"] or 30))
                    worker_amount = (line_total * worker_pct / Decimal("100")).quantize(Decimal("0.01"))
                    business_amount = line_total - worker_amount
                    subtotal += line_total
                    service_lines.append((service, worker_name, qty, price, line_total, worker_pct, business_pct, worker_amount, business_amount))

                tax_rate = Decimal(str(settings["tax_rate"] or 0))
                tax = (subtotal * tax_rate / Decimal("100")).quantize(Decimal("0.01"))
                total = subtotal + tax
                issued_date = request.form.get(
                    "issued_date",
                    ""
                ).strip()

                if issued_date:
                    issued_at = f"{issued_date}T12:00:00"
                else:
                    issued_at = datetime.now().isoformat(
                        timespec="seconds"
                    )
                    
                customer_id = None
                if customer_name.lower() != "consumidor final" or customer_document:
                    existing = None
                    if customer_document:
                        existing = conn.execute("SELECT id FROM customers WHERE document=? LIMIT 1", (customer_document,)).fetchone()
                    if existing:
                        customer_id = existing["id"]
                        conn.execute("UPDATE customers SET name=?,phone=?,email=?,address=? WHERE id=?",
                                     (customer_name, customer_phone, request.form.get("customer_email", ""), customer_address, customer_id))
                    else:
                        cur = conn.execute("INSERT INTO customers(document,name,phone,email,address) VALUES(?,?,?,?,?)",
                                           (customer_document, customer_name, customer_phone, request.form.get("customer_email", ""), customer_address))
                        customer_id = cur.lastrowid

                cur = conn.execute(
                    """INSERT INTO invoices(number,customer_id,customer_name,customer_document,customer_phone,
                    customer_address,vehicle_plate,salesperson_id,salesperson_name,issued_at,subtotal,tax,total,payment_method,notes,user_id)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (number, customer_id, customer_name, customer_document, customer_phone, customer_address, vehicle_plate,
                     salesperson["id"], salesperson["name"], issued_at, float(subtotal), float(tax), float(total),
                     payment_method, notes, g.user["id"]),
                )
                invoice_id = cur.lastrowid

                for product, qty, price, line_total in product_lines:
                    conn.execute(
                        """INSERT INTO invoice_items(invoice_id,product_id,code,description,quantity,unit_price,line_total)
                        VALUES(?,?,?,?,?,?,?)""",
                        (invoice_id, product["id"], product["code"], product["description"], float(qty), float(price), float(line_total)),
                    )
                    conn.execute("UPDATE products SET stock=stock-?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (float(qty), product["id"]))
                    conn.execute(
                        "INSERT INTO movements(product_id,movement_type,quantity,reference,notes,created_at,user_id) VALUES(?,?,?,?,?,?,?)",
                        (product["id"], "Venta", float(qty), number, "Venta registrada en remisión", issued_at, g.user["id"]),
                    )

                for service, worker_name, qty, price, line_total, worker_pct, business_pct, worker_amount, business_amount in service_lines:
                    conn.execute(
                        """INSERT INTO invoice_service_items(
                            invoice_id,service_id,service_name,worker_name,quantity,unit_price,line_total,
                            worker_percentage,business_percentage,worker_amount,business_amount)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                        (invoice_id, service["id"], service["name"], worker_name, float(qty), float(price), float(line_total),
                         float(worker_pct), float(business_pct), float(worker_amount), float(business_amount)),
                    )

                # Si la remisión es a crédito, crear automáticamente la cuenta por cobrar.
                if payment_method == "Crédito":
                    if credit_initial_payment > total:
                        raise ValueError(
                            "El abono inicial no puede ser mayor al total de la remisión."
                        )

                    initial_status = "PAGO PARCIAL" if credit_initial_payment > 0 else "PENDIENTE"
                    receivable_cur = conn.execute(
                        """INSERT INTO accounts_receivable(
                            invoice_id,original_amount,due_date,status,created_at
                        ) VALUES(?,?,?,?,?)""",
                        (
                            invoice_id,
                            float(total),
                            credit_due_date,
                            initial_status,
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    receivable_id = receivable_cur.lastrowid

                    if credit_initial_payment > 0:
                        conn.execute(
                            """INSERT INTO accounts_receivable_payments(
                                receivable_id,amount,payment_date,payment_method,notes,user_id
                            ) VALUES(?,?,?,?,?,?)""",
                            (
                                receivable_id,
                                float(credit_initial_payment),
                                datetime.now().date(),
                                credit_initial_payment_method,
                                "Abono inicial registrado al crear la remisión",
                                g.user["id"],
                            ),
                        )

                conn.execute("UPDATE settings SET next_invoice=next_invoice+1 WHERE id=1")

            generate_invoice_pdf(invoice_id)
            flash(f"Remisión {number} creada correctamente. El inventario de productos fue actualizado.", "success")
            return redirect(url_for("invoice_view", invoice_id=invoice_id))
        except (ValueError, KeyError) as exc:
            flash(str(exc), "danger")

    with db_conn() as conn:
        product_rows = conn.execute(
            "SELECT id,code,description,stock,sale_price,unit FROM products WHERE active=1 AND stock>0 ORDER BY description"
        ).fetchall()
        service_rows = conn.execute(
            "SELECT * FROM services WHERE active=1 ORDER BY name"
        ).fetchall()
        salesperson_rows = conn.execute(
            "SELECT id,name FROM salespeople WHERE active=1 ORDER BY name"
        ).fetchall()
        settings = get_settings(conn)
    return render_template(
        "invoice_form.html",
        products=product_rows,
        services=service_rows,
        salespeople=salesperson_rows,
        settings=settings,
        form_data=form_data
    )

@app.route("/finanzas")
@role_required("administrador")
def finances():
    """Resumen de ingresos registrados por productos y participación del negocio en servicios."""
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    # Validar fechas para evitar filtros inválidos y conservar consultas portables SQLite/PostgreSQL.
    for value, label in ((date_from, "Desde"), (date_to, "Hasta")):
        if value:
            try:
                datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                flash(f"La fecha '{label}' no es válida.", "danger")
                return redirect(url_for("finances"))

    if date_from and date_to and date_from > date_to:
        flash("La fecha Desde no puede ser posterior a la fecha Hasta.", "danger")
        return redirect(url_for("finances"))

    filters = []
    params = []
    if date_from:
        filters.append("DATE(i.issued_at) >= ?")
        params.append(date_from)
    if date_to:
        filters.append("DATE(i.issued_at) <= ?")
        params.append(date_to)

    where_clause = ""
    if filters:
        where_clause = " WHERE " + " AND ".join(filters)

    with db_conn() as conn:
        product_summary = conn.execute(
            f"""
            SELECT
                COALESCE(SUM(ii.line_total), 0) AS product_sales,
                COALESCE(SUM(ii.quantity), 0) AS product_units,
                COUNT(ii.id) AS product_lines
            FROM invoice_items ii
            JOIN invoices i ON i.id = ii.invoice_id
            {where_clause}
            """,
            tuple(params),
        ).fetchone()

        service_summary = conn.execute(
            f"""
            SELECT
                COALESCE(SUM(isi.line_total), 0) AS service_total,
                COALESCE(SUM(isi.worker_amount), 0) AS worker_total,
                COALESCE(SUM(isi.business_amount), 0) AS business_total,
                COUNT(isi.id) AS service_lines
            FROM invoice_service_items isi
            JOIN invoices i ON i.id = isi.invoice_id
            {where_clause}
            """,
            tuple(params),
        ).fetchone()

        invoice_summary = conn.execute(
            f"""
            SELECT COUNT(*) AS invoice_count
            FROM invoices i
            {where_clause}
            """,
            tuple(params),
        ).fetchone()

        daily = conn.execute(
            f"""
            SELECT
                activity_date,
                COALESCE(SUM(product_sales), 0) AS product_sales,
                COALESCE(SUM(service_total), 0) AS service_total,
                COALESCE(SUM(service_business), 0) AS service_business
            FROM (
                SELECT
                    DATE(i.issued_at) AS activity_date,
                    COALESCE(SUM(ii.line_total), 0) AS product_sales,
                    0 AS service_total,
                    0 AS service_business
                FROM invoice_items ii
                JOIN invoices i ON i.id = ii.invoice_id
                {where_clause}
                GROUP BY DATE(i.issued_at)

                UNION ALL

                SELECT
                    DATE(i.issued_at) AS activity_date,
                    0 AS product_sales,
                    COALESCE(SUM(isi.line_total), 0) AS service_total,
                    COALESCE(SUM(isi.business_amount), 0) AS service_business
                FROM invoice_service_items isi
                JOIN invoices i ON i.id = isi.invoice_id
                {where_clause}
                GROUP BY DATE(i.issued_at)
            ) finance_daily
            GROUP BY activity_date
            ORDER BY activity_date DESC
            """,
            tuple(params) + tuple(params),
        ).fetchall()

        product_detail = conn.execute(
            f"""
            SELECT
                ii.code,
                ii.description,
                COALESCE(SUM(ii.quantity), 0) AS quantity_total,
                COALESCE(SUM(ii.line_total), 0) AS sales_total
            FROM invoice_items ii
            JOIN invoices i ON i.id = ii.invoice_id
            {where_clause}
            GROUP BY ii.code, ii.description
            ORDER BY sales_total DESC, ii.description ASC
            """,
            tuple(params),
        ).fetchall()

    product_sales = Decimal(str(product_summary["product_sales"] or 0))
    service_total = Decimal(str(service_summary["service_total"] or 0))
    service_worker = Decimal(str(service_summary["worker_total"] or 0))
    service_business = Decimal(str(service_summary["business_total"] or 0))
    business_income = product_sales + service_business

    daily_rows = []
    for row in daily:
        item = dict(row)
        item["business_income"] = (
            Decimal(str(item.get("product_sales") or 0))
            + Decimal(str(item.get("service_business") or 0))
        )
        daily_rows.append(item)

    now = datetime.now().date()
    yesterday = now.fromordinal(now.toordinal() - 1)
    week_start = now.fromordinal(now.toordinal() - now.weekday())
    month_start = now.replace(day=1)

    return render_template(
        "finances.html",
        date_from=date_from,
        date_to=date_to,
        product_sales=product_sales,
        service_total=service_total,
        service_worker=service_worker,
        service_business=service_business,
        business_income=business_income,
        invoice_count=int(invoice_summary["invoice_count"] or 0),
        product_units=Decimal(str(product_summary["product_units"] or 0)),
        daily=daily_rows,
        product_detail=product_detail,
        today_iso=now.isoformat(),
        yesterday_iso=yesterday.isoformat(),
        week_start_iso=week_start.isoformat(),
        month_start_iso=month_start.isoformat(),
    )


def _receivable_display_status(row: dict[str, Any], today=None) -> str:
    """Calcula el estado visible de una cuenta por cobrar."""
    today = today or datetime.now().date()
    original = Decimal(str(row.get("original_amount") or 0))
    paid = Decimal(str(row.get("paid_amount") or 0))
    balance = original - paid

    if balance <= 0:
        return "PAGADA"

    due_raw = row.get("due_date")
    if due_raw:
        if hasattr(due_raw, "date") and not isinstance(due_raw, str):
            try:
                due = due_raw.date()
            except Exception:
                due = due_raw
        elif isinstance(due_raw, str):
            try:
                due = datetime.strptime(due_raw[:10], "%Y-%m-%d").date()
            except ValueError:
                due = None
        else:
            due = due_raw
        if due and due < today:
            return "VENCIDA"

    if paid > 0:
        return "PAGO PARCIAL"
    return "PENDIENTE"


@app.route("/cuentas-por-cobrar")
@login_required
def accounts_receivable():
    status_filter = request.args.get("status", "ABIERTAS").strip().upper()
    q = request.args.get("q", "").strip()

    with db_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                ar.id, ar.invoice_id, ar.original_amount, ar.due_date,
                ar.status, ar.created_at, ar.closed_at,
                i.number, i.issued_at, i.customer_name, i.customer_document,
                i.customer_phone, i.vehicle_plate,
                COALESCE((
                    SELECT SUM(p.amount)
                    FROM accounts_receivable_payments p
                    WHERE p.receivable_id = ar.id
                ), 0) AS paid_amount
            FROM accounts_receivable ar
            JOIN invoices i ON i.id = ar.invoice_id
            ORDER BY i.issued_at ASC, ar.id ASC
            """
        ).fetchall()

    accounts = []
    total_pending = Decimal("0")
    total_overdue = Decimal("0")
    total_without_date = Decimal("0")

    q_lower = q.lower()
    for row in rows:
        account = dict(row)
        original = Decimal(str(account.get("original_amount") or 0))
        paid = Decimal(str(account.get("paid_amount") or 0))
        balance = max(Decimal("0"), original - paid)
        display_status = _receivable_display_status(account)

        account["paid_amount"] = paid
        account["balance"] = balance
        account["display_status"] = display_status

        if balance > 0:
            total_pending += balance
            if display_status == "VENCIDA":
                total_overdue += balance
            if not account.get("due_date"):
                total_without_date += balance

        if q_lower:
            haystack = " ".join(
                str(account.get(key) or "")
                for key in ("number", "customer_name", "customer_document", "customer_phone", "vehicle_plate")
            ).lower()
            if q_lower not in haystack:
                continue

        if status_filter == "ABIERTAS" and display_status == "PAGADA":
            continue
        if status_filter not in ("", "TODAS", "ABIERTAS") and display_status != status_filter:
            continue

        accounts.append(account)

    return render_template(
        "accounts_receivable.html",
        accounts=accounts,
        status_filter=status_filter,
        q=q,
        total_pending=total_pending,
        total_overdue=total_overdue,
        total_without_date=total_without_date,
    )


@app.route("/cuentas-por-cobrar/<int:account_id>/pago", methods=["POST"])
@login_required
def accounts_receivable_payment(account_id: int):
    try:
        amount = parse_decimal(request.form.get("amount", "0"))
        payment_method = request.form.get("payment_method", "Efectivo").strip() or "Efectivo"
        payment_date_raw = request.form.get("payment_date", "").strip()
        notes = request.form.get("notes", "").strip()

        if amount <= 0:
            raise ValueError("El valor del pago debe ser mayor a cero.")

        if payment_date_raw:
            try:
                payment_date = datetime.strptime(payment_date_raw, "%Y-%m-%d").date()
            except ValueError as exc:
                raise ValueError("La fecha del pago no es válida.") from exc
        else:
            payment_date = datetime.now().date()

        with db_conn() as conn:
            account = conn.execute(
                """
                SELECT ar.*,
                       COALESCE((
                           SELECT SUM(p.amount)
                           FROM accounts_receivable_payments p
                           WHERE p.receivable_id = ar.id
                       ), 0) AS paid_amount
                FROM accounts_receivable ar
                WHERE ar.id=?
                """,
                (account_id,),
            ).fetchone()

            if not account:
                raise ValueError("La cuenta por cobrar no existe.")

            original = Decimal(str(account["original_amount"] or 0))
            already_paid = Decimal(str(account["paid_amount"] or 0))
            balance = original - already_paid

            if balance <= 0:
                raise ValueError("Esta cuenta ya se encuentra pagada.")
            if amount > balance:
                raise ValueError(f"El pago supera el saldo pendiente de {money(balance)}.")

            conn.execute(
                """INSERT INTO accounts_receivable_payments(
                    receivable_id,amount,payment_date,payment_method,notes,user_id
                ) VALUES(?,?,?,?,?,?)""",
                (account_id, float(amount), payment_date, payment_method, notes, g.user["id"]),
            )

            new_balance = balance - amount
            new_status = "PAGADA" if new_balance <= 0 else "PAGO PARCIAL"
            closed_at = datetime.now().isoformat(timespec="seconds") if new_balance <= 0 else None
            conn.execute(
                "UPDATE accounts_receivable SET status=?,closed_at=? WHERE id=?",
                (new_status, closed_at, account_id),
            )

        flash("Pago registrado correctamente.", "success")
    except ValueError as exc:
        flash(str(exc), "danger")

    return redirect(url_for("accounts_receivable"))


@app.route("/cuentas-por-cobrar/<int:account_id>/pagar-total", methods=["POST"])
@login_required
def accounts_receivable_pay_full(account_id: int):
    try:
        payment_method = request.form.get("payment_method", "Efectivo").strip() or "Efectivo"
        payment_date_raw = request.form.get("payment_date", "").strip()
        notes = request.form.get("notes", "").strip() or "Pago total de la cuenta"

        if payment_date_raw:
            try:
                payment_date = datetime.strptime(payment_date_raw, "%Y-%m-%d").date()
            except ValueError as exc:
                raise ValueError("La fecha del pago no es válida.") from exc
        else:
            payment_date = datetime.now().date()

        with db_conn() as conn:
            account = conn.execute(
                """
                SELECT ar.*,
                       COALESCE((
                           SELECT SUM(p.amount)
                           FROM accounts_receivable_payments p
                           WHERE p.receivable_id = ar.id
                       ), 0) AS paid_amount
                FROM accounts_receivable ar
                WHERE ar.id=?
                """,
                (account_id,),
            ).fetchone()

            if not account:
                raise ValueError("La cuenta por cobrar no existe.")

            original = Decimal(str(account["original_amount"] or 0))
            paid = Decimal(str(account["paid_amount"] or 0))
            balance = original - paid

            if balance > 0:
                conn.execute(
                    """INSERT INTO accounts_receivable_payments(
                        receivable_id,amount,payment_date,payment_method,notes,user_id
                    ) VALUES(?,?,?,?,?,?)""",
                    (account_id, float(balance), payment_date, payment_method, notes, g.user["id"]),
                )

            conn.execute(
                "UPDATE accounts_receivable SET status='PAGADA',closed_at=? WHERE id=?",
                (datetime.now().isoformat(timespec="seconds"), account_id),
            )

        flash("Cuenta marcada como pagada correctamente.", "success")
    except ValueError as exc:
        flash(str(exc), "danger")

    return redirect(url_for("accounts_receivable"))


@app.route("/cuentas-por-cobrar/<int:account_id>/historial")
@login_required
def accounts_receivable_history(account_id: int):
    with db_conn() as conn:
        account = conn.execute(
            """
            SELECT ar.*, i.number, i.issued_at, i.customer_name, i.customer_phone, i.vehicle_plate
            FROM accounts_receivable ar
            JOIN invoices i ON i.id=ar.invoice_id
            WHERE ar.id=?
            """,
            (account_id,),
        ).fetchone()
        payments = conn.execute(
            """
            SELECT p.*, u.full_name AS user_name
            FROM accounts_receivable_payments p
            LEFT JOIN users u ON u.id=p.user_id
            WHERE p.receivable_id=?
            ORDER BY p.payment_date DESC, p.id DESC
            """,
            (account_id,),
        ).fetchall()

    if not account:
        flash("Cuenta por cobrar no encontrada.", "danger")
        return redirect(url_for("accounts_receivable"))

    total_paid = sum((Decimal(str(p["amount"] or 0)) for p in payments), Decimal("0"))
    balance = max(Decimal("0"), Decimal(str(account["original_amount"] or 0)) - total_paid)
    return render_template(
        "accounts_receivable_history.html",
        account=account,
        payments=payments,
        total_paid=total_paid,
        balance=balance,
    )

# ============================================================
# CUENTAS POR PAGAR / FACTURAS PENDIENTES
# ============================================================


def _parse_payable_date(value):
    """Convierte distintos formatos devueltos por SQLite/PostgreSQL a date."""

    if not value:
        return None

    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except Exception:
            pass

    if isinstance(value, str):
        try:
            return datetime.strptime(
                value[:10],
                "%Y-%m-%d"
            ).date()
        except ValueError:
            return None

    return value


def _payable_display_status(row, today=None):
    """
    Estado visual tipo semáforo.

    PAGADA             -> pago completo
    VENCIDA            -> fecha límite ya pasó
    PROXIMA A VENCER   -> faltan entre 0 y 7 días
    PENDIENTE          -> faltan más de 7 días
    """

    today = today or datetime.now().date()

    original_amount = Decimal(
        str(
            row.get(
                "original_amount"
            )
            or
            0
        )
    )

    paid_amount = Decimal(
        str(
            row.get(
                "paid_amount"
            )
            or
            0
        )
    )

    balance = (
        original_amount
        -
        paid_amount
    )

    if balance <= Decimal("0"):
        return "PAGADA"

    due_date = _parse_payable_date(
        row.get(
            "due_date"
        )
    )

    if not due_date:
        return "PENDIENTE"

    days_remaining = (
        due_date
        -
        today
    ).days

    if days_remaining < 0:
        return "VENCIDA"

    if days_remaining <= 7:
        return "PROXIMA A VENCER"

    return "PENDIENTE"


def _payable_days_remaining(row, today=None):

    today = today or datetime.now().date()

    due_date = _parse_payable_date(
        row.get(
            "due_date"
        )
    )

    if not due_date:
        return None

    return (
        due_date
        -
        today
    ).days


# ============================================================
# LISTADO
# ============================================================

@app.route("/facturas-pendientes")
@role_required("administrador")
def accounts_payable():

    status_filter = (
        request.args.get(
            "status",
            "ABIERTAS"
        )
        .strip()
        .upper()
    )

    q = (
        request.args.get(
            "q",
            ""
        )
        .strip()
    )

    with db_conn() as conn:

        rows = conn.execute(
            """
            SELECT
                ap.*,

                COALESCE(
                    (
                        SELECT SUM(p.amount)
                        FROM accounts_payable_payments p
                        WHERE p.payable_id = ap.id
                    ),
                    0
                ) AS paid_amount

            FROM accounts_payable ap

            ORDER BY
                ap.due_date ASC,
                ap.id DESC
            """
        ).fetchall()


    accounts = []

    total_open = Decimal("0")
    total_pending = Decimal("0")
    total_upcoming = Decimal("0")
    total_overdue = Decimal("0")

    count_pending = 0
    count_upcoming = 0
    count_overdue = 0

    q_lower = q.lower()


    for row in rows:

        account = dict(row)

        original_amount = Decimal(
            str(
                account.get(
                    "original_amount"
                )
                or
                0
            )
        )

        paid_amount = Decimal(
            str(
                account.get(
                    "paid_amount"
                )
                or
                0
            )
        )

        balance = max(
            Decimal("0"),
            original_amount
            -
            paid_amount
        )

        display_status = (
            _payable_display_status(
                account
            )
        )

        days_remaining = (
            _payable_days_remaining(
                account
            )
        )


        account["paid_amount"] = (
            paid_amount
        )

        account["balance"] = (
            balance
        )

        account["display_status"] = (
            display_status
        )

        account["days_remaining"] = (
            days_remaining
        )


        if balance > 0:

            total_open += balance

            if display_status == "PENDIENTE":

                total_pending += balance
                count_pending += 1

            elif display_status == "PROXIMA A VENCER":

                total_upcoming += balance
                count_upcoming += 1

            elif display_status == "VENCIDA":

                total_overdue += balance
                count_overdue += 1


        # Buscador
        if q_lower:

            searchable = " ".join(
                str(
                    account.get(key)
                    or
                    ""
                )
                for key in (
                    "supplier",
                    "invoice_number",
                    "description",
                    "notes",
                )
            ).lower()

            if q_lower not in searchable:
                continue


        # Filtros
        if (
            status_filter == "ABIERTAS"
            and
            display_status == "PAGADA"
        ):
            continue


        if (
            status_filter
            not in (
                "",
                "TODAS",
                "ABIERTAS"
            )
            and
            display_status
            !=
            status_filter
        ):
            continue


        accounts.append(
            account
        )


    return render_template(
        "accounts_payable.html",

        accounts=accounts,

        status_filter=status_filter,

        q=q,

        total_open=total_open,

        total_pending=total_pending,

        total_upcoming=total_upcoming,

        total_overdue=total_overdue,

        count_pending=count_pending,

        count_upcoming=count_upcoming,

        count_overdue=count_overdue,
    )


# ============================================================
# CREAR FACTURA
# ============================================================

@app.route(
    "/facturas-pendientes/nueva",
    methods=["GET", "POST"]
)
@role_required("administrador")
def account_payable_new():

    if request.method == "POST":

        try:

            supplier = (
                request.form.get(
                    "supplier",
                    ""
                )
                .strip()
            )

            invoice_number = (
                request.form.get(
                    "invoice_number",
                    ""
                )
                .strip()
            )

            purchase_date_raw = (
                request.form.get(
                    "purchase_date",
                    ""
                )
                .strip()
            )

            due_date_raw = (
                request.form.get(
                    "due_date",
                    ""
                )
                .strip()
            )

            amount = parse_decimal(
                request.form.get(
                    "original_amount",
                    "0"
                )
            )

            description = (
                request.form.get(
                    "description",
                    ""
                )
                .strip()
            )

            notes = (
                request.form.get(
                    "notes",
                    ""
                )
                .strip()
            )


            if not supplier:

                raise ValueError(
                    "El proveedor es obligatorio."
                )


            if not invoice_number:

                raise ValueError(
                    "El número de factura es obligatorio."
                )


            if not purchase_date_raw:

                raise ValueError(
                    "La fecha de compra es obligatoria."
                )


            if not due_date_raw:

                raise ValueError(
                    "La fecha límite de pago es obligatoria."
                )


            try:

                purchase_date = datetime.strptime(
                    purchase_date_raw,
                    "%Y-%m-%d"
                ).date()

            except ValueError as exc:

                raise ValueError(
                    "La fecha de compra no es válida."
                ) from exc


            try:

                due_date = datetime.strptime(
                    due_date_raw,
                    "%Y-%m-%d"
                ).date()

            except ValueError as exc:

                raise ValueError(
                    "La fecha límite de pago no es válida."
                ) from exc


            if due_date < purchase_date:

                raise ValueError(
                    "La fecha límite de pago no puede ser anterior a la fecha de compra."
                )


            if amount <= Decimal("0"):

                raise ValueError(
                    "El valor de la factura debe ser mayor a cero."
                )


            with db_conn() as conn:

                duplicate = conn.execute(
                    """
                    SELECT id
                    FROM accounts_payable
                    WHERE LOWER(supplier)=LOWER(?)
                    AND LOWER(invoice_number)=LOWER(?)
                    """,
                    (
                        supplier,
                        invoice_number
                    )
                ).fetchone()


                if duplicate:

                    raise ValueError(
                        "Ya existe una factura con ese número para este proveedor."
                    )


                conn.execute(
                    """
                    INSERT INTO accounts_payable(
                        supplier,
                        invoice_number,
                        purchase_date,
                        due_date,
                        original_amount,
                        description,
                        notes,
                        status,
                        user_id,
                        created_at,
                        updated_at
                    )
                    VALUES(
                        ?,?,?,?,?,?,?,?,
                        ?,?,?
                    )
                    """,
                    (
                        supplier,
                        invoice_number,
                        purchase_date,
                        due_date,
                        float(amount),
                        description,
                        notes,
                        "PENDIENTE",
                        g.user["id"],
                        datetime.now().isoformat(
                            timespec="seconds"
                        ),
                        datetime.now().isoformat(
                            timespec="seconds"
                        ),
                    )
                )


            flash(
                "Factura pendiente registrada correctamente.",
                "success"
            )

            return redirect(
                url_for(
                    "accounts_payable"
                )
            )


        except ValueError as exc:

            flash(
                str(exc),
                "danger"
            )


    return render_template(
        "account_payable_form.html",
        account=None
    )


# ============================================================
# EDITAR FACTURA
# ============================================================

@app.route(
    "/facturas-pendientes/<int:account_id>/editar",
    methods=["GET", "POST"]
)
@role_required("administrador")
def account_payable_edit(account_id):

    with db_conn() as conn:

        account = conn.execute(
            """
            SELECT ap.*,

                   COALESCE(
                       (
                           SELECT SUM(p.amount)
                           FROM accounts_payable_payments p
                           WHERE p.payable_id=ap.id
                       ),
                       0
                   ) AS paid_amount

            FROM accounts_payable ap

            WHERE ap.id=?
            """,
            (
                account_id,
            )
        ).fetchone()


    if not account:

        flash(
            "Factura pendiente no encontrada.",
            "danger"
        )

        return redirect(
            url_for(
                "accounts_payable"
            )
        )


    if request.method == "POST":

        try:

            supplier = (
                request.form.get(
                    "supplier",
                    ""
                )
                .strip()
            )

            invoice_number = (
                request.form.get(
                    "invoice_number",
                    ""
                )
                .strip()
            )

            purchase_date_raw = (
                request.form.get(
                    "purchase_date",
                    ""
                )
                .strip()
            )

            due_date_raw = (
                request.form.get(
                    "due_date",
                    ""
                )
                .strip()
            )

            amount = parse_decimal(
                request.form.get(
                    "original_amount",
                    "0"
                )
            )

            description = (
                request.form.get(
                    "description",
                    ""
                )
                .strip()
            )

            notes = (
                request.form.get(
                    "notes",
                    ""
                )
                .strip()
            )


            if not supplier:
                raise ValueError(
                    "El proveedor es obligatorio."
                )


            if not invoice_number:
                raise ValueError(
                    "El número de factura es obligatorio."
                )


            try:

                purchase_date = datetime.strptime(
                    purchase_date_raw,
                    "%Y-%m-%d"
                ).date()

                due_date = datetime.strptime(
                    due_date_raw,
                    "%Y-%m-%d"
                ).date()

            except ValueError as exc:

                raise ValueError(
                    "Las fechas ingresadas no son válidas."
                ) from exc


            if due_date < purchase_date:

                raise ValueError(
                    "La fecha límite no puede ser anterior a la fecha de compra."
                )


            if amount <= 0:

                raise ValueError(
                    "El valor de la factura debe ser mayor a cero."
                )


            paid_amount = Decimal(
                str(
                    account.get(
                        "paid_amount"
                    )
                    or
                    0
                )
            )


            if amount < paid_amount:

                raise ValueError(
                    "El valor de la factura no puede ser menor al valor que ya fue pagado."
                )


            with db_conn() as conn:

                duplicate = conn.execute(
                    """
                    SELECT id
                    FROM accounts_payable

                    WHERE LOWER(supplier)=LOWER(?)
                    AND LOWER(invoice_number)=LOWER(?)
                    AND id<>?
                    """,
                    (
                        supplier,
                        invoice_number,
                        account_id
                    )
                ).fetchone()


                if duplicate:

                    raise ValueError(
                        "Ya existe otra factura con ese número para este proveedor."
                    )


                balance = (
                    amount
                    -
                    paid_amount
                )


                new_status = (
                    "PAGADA"
                    if balance <= 0
                    else
                    (
                        "PAGO PARCIAL"
                        if paid_amount > 0
                        else
                        "PENDIENTE"
                    )
                )


                closed_at = (
                    datetime.now().isoformat(
                        timespec="seconds"
                    )
                    if new_status == "PAGADA"
                    else
                    None
                )


                conn.execute(
                    """
                    UPDATE accounts_payable

                    SET
                        supplier=?,
                        invoice_number=?,
                        purchase_date=?,
                        due_date=?,
                        original_amount=?,
                        description=?,
                        notes=?,
                        status=?,
                        closed_at=?,
                        updated_at=CURRENT_TIMESTAMP

                    WHERE id=?
                    """,
                    (
                        supplier,
                        invoice_number,
                        purchase_date,
                        due_date,
                        float(amount),
                        description,
                        notes,
                        new_status,
                        closed_at,
                        account_id
                    )
                )


            flash(
                "Factura actualizada correctamente.",
                "success"
            )

            return redirect(
                url_for(
                    "accounts_payable"
                )
            )


        except ValueError as exc:

            flash(
                str(exc),
                "danger"
            )


    return render_template(
        "account_payable_form.html",
        account=account
    )


# ============================================================
# PAGO PARCIAL
# ============================================================

@app.route(
    "/facturas-pendientes/<int:account_id>/pago",
    methods=["POST"]
)
@role_required("administrador")
def account_payable_payment(account_id):

    try:

        amount = parse_decimal(
            request.form.get(
                "amount",
                "0"
            )
        )

        payment_method = (
            request.form.get(
                "payment_method",
                "Transferencia"
            )
            .strip()
            or
            "Transferencia"
        )

        payment_date_raw = (
            request.form.get(
                "payment_date",
                ""
            )
            .strip()
        )

        notes = (
            request.form.get(
                "notes",
                ""
            )
            .strip()
        )


        if amount <= 0:

            raise ValueError(
                "El valor del pago debe ser mayor a cero."
            )


        if payment_date_raw:

            try:

                payment_date = datetime.strptime(
                    payment_date_raw,
                    "%Y-%m-%d"
                ).date()

            except ValueError as exc:

                raise ValueError(
                    "La fecha del pago no es válida."
                ) from exc

        else:

            payment_date = (
                datetime.now().date()
            )


        with db_conn() as conn:

            account = conn.execute(
                """
                SELECT
                    ap.*,

                    COALESCE(
                        (
                            SELECT SUM(p.amount)
                            FROM accounts_payable_payments p
                            WHERE p.payable_id=ap.id
                        ),
                        0
                    ) AS paid_amount

                FROM accounts_payable ap

                WHERE ap.id=?
                """,
                (
                    account_id,
                )
            ).fetchone()


            if not account:

                raise ValueError(
                    "La factura pendiente no existe."
                )


            original = Decimal(
                str(
                    account["original_amount"]
                    or
                    0
                )
            )

            already_paid = Decimal(
                str(
                    account["paid_amount"]
                    or
                    0
                )
            )

            balance = (
                original
                -
                already_paid
            )


            if balance <= 0:

                raise ValueError(
                    "Esta factura ya se encuentra pagada."
                )


            if amount > balance:

                raise ValueError(
                    f"El pago supera el saldo pendiente de {money(balance)}."
                )


            conn.execute(
                """
                INSERT INTO accounts_payable_payments(
                    payable_id,
                    amount,
                    payment_date,
                    payment_method,
                    notes,
                    user_id
                )
                VALUES(?,?,?,?,?,?)
                """,
                (
                    account_id,
                    float(amount),
                    payment_date,
                    payment_method,
                    notes,
                    g.user["id"]
                )
            )


            new_balance = (
                balance
                -
                amount
            )


            new_status = (
                "PAGADA"
                if new_balance <= 0
                else
                "PAGO PARCIAL"
            )


            closed_at = (
                datetime.now().isoformat(
                    timespec="seconds"
                )
                if new_balance <= 0
                else
                None
            )


            conn.execute(
                """
                UPDATE accounts_payable
                SET
                    status=?,
                    closed_at=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=?
                """,
                (
                    new_status,
                    closed_at,
                    account_id
                )
            )


        flash(
            "Pago registrado correctamente.",
            "success"
        )


    except ValueError as exc:

        flash(
            str(exc),
            "danger"
        )


    return redirect(
        url_for(
            "accounts_payable"
        )
    )


# ============================================================
# PAGAR TODO
# ============================================================

@app.route(
    "/facturas-pendientes/<int:account_id>/pagar-total",
    methods=["POST"]
)
@role_required("administrador")
def account_payable_pay_full(account_id):

    try:

        payment_method = (
            request.form.get(
                "payment_method",
                "Transferencia"
            )
            .strip()
            or
            "Transferencia"
        )

        payment_date_raw = (
            request.form.get(
                "payment_date",
                ""
            )
            .strip()
        )

        notes = (
            request.form.get(
                "notes",
                ""
            )
            .strip()
            or
            "Pago total de factura"
        )


        if payment_date_raw:

            payment_date = datetime.strptime(
                payment_date_raw,
                "%Y-%m-%d"
            ).date()

        else:

            payment_date = (
                datetime.now().date()
            )


        with db_conn() as conn:

            account = conn.execute(
                """
                SELECT
                    ap.*,

                    COALESCE(
                        (
                            SELECT SUM(p.amount)
                            FROM accounts_payable_payments p
                            WHERE p.payable_id=ap.id
                        ),
                        0
                    ) AS paid_amount

                FROM accounts_payable ap

                WHERE ap.id=?
                """,
                (
                    account_id,
                )
            ).fetchone()


            if not account:

                raise ValueError(
                    "Factura pendiente no encontrada."
                )


            original = Decimal(
                str(
                    account["original_amount"]
                    or
                    0
                )
            )

            already_paid = Decimal(
                str(
                    account["paid_amount"]
                    or
                    0
                )
            )

            balance = (
                original
                -
                already_paid
            )


            if balance > 0:

                conn.execute(
                    """
                    INSERT INTO accounts_payable_payments(
                        payable_id,
                        amount,
                        payment_date,
                        payment_method,
                        notes,
                        user_id
                    )
                    VALUES(?,?,?,?,?,?)
                    """,
                    (
                        account_id,
                        float(balance),
                        payment_date,
                        payment_method,
                        notes,
                        g.user["id"]
                    )
                )


            conn.execute(
                """
                UPDATE accounts_payable

                SET
                    status='PAGADA',
                    closed_at=?,
                    updated_at=CURRENT_TIMESTAMP

                WHERE id=?
                """,
                (
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                    account_id
                )
            )


        flash(
            "Factura marcada como pagada.",
            "success"
        )


    except ValueError as exc:

        flash(
            str(exc),
            "danger"
        )


    return redirect(
        url_for(
            "accounts_payable"
        )
    )


# ============================================================
# HISTORIAL
# ============================================================

@app.route(
    "/facturas-pendientes/<int:account_id>/historial"
)
@role_required("administrador")
def account_payable_history(account_id):

    with db_conn() as conn:

        account = conn.execute(
            """
            SELECT ap.*,

                   COALESCE(
                       (
                           SELECT SUM(p.amount)
                           FROM accounts_payable_payments p
                           WHERE p.payable_id=ap.id
                       ),
                       0
                   ) AS paid_amount

            FROM accounts_payable ap

            WHERE ap.id=?
            """,
            (
                account_id,
            )
        ).fetchone()


        payments = conn.execute(
            """
            SELECT
                p.*,
                u.full_name AS user_name

            FROM accounts_payable_payments p

            LEFT JOIN users u
                ON u.id=p.user_id

            WHERE p.payable_id=?

            ORDER BY
                p.payment_date DESC,
                p.id DESC
            """,
            (
                account_id,
            )
        ).fetchall()


    if not account:

        flash(
            "Factura pendiente no encontrada.",
            "danger"
        )

        return redirect(
            url_for(
                "accounts_payable"
            )
        )


    account = dict(
        account
    )


    original = Decimal(
        str(
            account.get(
                "original_amount"
            )
            or
            0
        )
    )


    paid = Decimal(
        str(
            account.get(
                "paid_amount"
            )
            or
            0
        )
    )


    balance = max(
        Decimal("0"),
        original
        -
        paid
    )


    account["balance"] = (
        balance
    )

    account["display_status"] = (
        _payable_display_status(
            account
        )
    )


    return render_template(
        "account_payable_history.html",
        account=account,
        payments=payments,
        total_paid=paid,
        balance=balance,
    )


@app.route("/facturas/<int:invoice_id>")
@login_required
def invoice_view(invoice_id: int):

    with db_conn() as conn:

        invoice = conn.execute(
            """
            SELECT
                i.*,
                u.full_name AS user_name
            FROM invoices i

            LEFT JOIN users u
                ON u.id=i.user_id

            WHERE i.id=?
            """,
            (invoice_id,)
        ).fetchone()

        items = conn.execute(
            """
            SELECT *
            FROM invoice_items
            WHERE invoice_id=?
            """,
            (invoice_id,)
        ).fetchall()

        service_items = conn.execute(
            """
            SELECT *
            FROM invoice_service_items
            WHERE invoice_id=?
            """,
            (invoice_id,)
        ).fetchall()

        settings = get_settings(conn)


    if not invoice:

        flash(
            "Remisión no encontrada.",
            "danger"
        )

        return redirect(
            url_for("invoices")
        )


    # =====================================================
    # PERMISO PARA EDITAR REMISIONES
    # =====================================================

    current_username = str(
        g.user["username"] or ""
    ).strip().lower()

    current_full_name = str(
        g.user["full_name"] or ""
    ).strip().lower()

    current_role = str(
        g.user["role"] or ""
    ).strip().lower()


    can_edit_invoice = (
        current_role == "administrador"
        and (
            current_username == "arturo.lopez"
            or current_full_name in (
                "arturo lópez",
                "arturo lopez"
            )
        )
    )


    return render_template(
        "invoice_view.html",

        invoice=invoice,
        items=items,
        service_items=service_items,
        settings=settings,

        can_edit_invoice=can_edit_invoice
    )


@app.route("/facturas/<int:invoice_id>/editar", methods=["GET", "POST"])
@arturo_required
def invoice_edit(invoice_id: int):
    """Edita una remisión y ajusta inventario/cartera de forma transaccional."""

    with db_conn() as conn:
        invoice = conn.execute(
            "SELECT * FROM invoices WHERE id=?",
            (invoice_id,),
        ).fetchone()

    if not invoice:
        flash("Remisión no encontrada.", "danger")
        return redirect(url_for("invoices"))

    if request.method == "POST":
        try:
            customer_name = (
                request.form.get("customer_name", "Consumidor final").strip()
                or "Consumidor final"
            )
            customer_document = request.form.get("customer_document", "").strip()
            customer_phone = request.form.get("customer_phone", "").strip()
            customer_address = request.form.get("customer_address", "").strip()
            customer_email = request.form.get("customer_email", "").strip()
            vehicle_plate = request.form.get("vehicle_plate", "").strip().upper()
            salesperson_id_raw = request.form.get("salesperson_id", "").strip()
            payment_method = request.form.get("payment_method", "Efectivo").strip()
            notes = request.form.get("notes", "").strip()
            issued_date = request.form.get("issued_date", "").strip()
            credit_due_date = request.form.get("credit_due_date", "").strip()

            if not salesperson_id_raw:
                raise ValueError("Selecciona quién realizó la venta.")
            salesperson_id = int(salesperson_id_raw)

            if payment_method not in ("Efectivo", "Transferencia", "Tarjeta", "Crédito"):
                raise ValueError("Forma de pago inválida.")

            if issued_date:
                try:
                    datetime.strptime(issued_date, "%Y-%m-%d")
                except ValueError as exc:
                    raise ValueError("La fecha de la remisión no es válida.") from exc
                issued_at = f"{issued_date}T12:00:00"
            else:
                issued_at = invoice["issued_at"]

            if credit_due_date:
                try:
                    datetime.strptime(credit_due_date, "%Y-%m-%d")
                except ValueError as exc:
                    raise ValueError("La fecha acordada de pago no es válida.") from exc

            product_rows_raw = [
                x
                for x in zip(
                    request.form.getlist("product_id[]"),
                    request.form.getlist("quantity[]"),
                    request.form.getlist("price[]"),
                )
                if str(x[0]).strip()
            ]

            service_rows_raw = [
                x
                for x in zip(
                    request.form.getlist("service_id[]"),
                    request.form.getlist("service_worker[]"),
                    request.form.getlist("service_quantity[]"),
                    request.form.getlist("service_price[]"),
                )
                if str(x[0]).strip()
            ]

            if not product_rows_raw and not service_rows_raw:
                raise ValueError(
                    "La remisión debe contener al menos un producto o un servicio."
                )

            with db_conn() as conn:
                invoice_sql = "SELECT * FROM invoices WHERE id=?"
                if db.engine.dialect.name == "postgresql":
                    invoice_sql += " FOR UPDATE"
                current_invoice = conn.execute(invoice_sql, (invoice_id,)).fetchone()
                if not current_invoice:
                    raise ValueError("La remisión ya no existe.")

                salesperson = conn.execute(
                    "SELECT id,name FROM salespeople WHERE id=? AND active=1",
                    (salesperson_id,),
                ).fetchone()
                if not salesperson:
                    raise ValueError("El vendedor seleccionado no está disponible.")

                # Cantidades anteriores por producto.
                old_items = conn.execute(
                    "SELECT * FROM invoice_items WHERE invoice_id=?",
                    (invoice_id,),
                ).fetchall()
                old_quantities: dict[int, Decimal] = {}
                for item in old_items:
                    pid = int(item["product_id"])
                    old_quantities[pid] = old_quantities.get(pid, Decimal("0")) + Decimal(
                        str(item["quantity"] or 0)
                    )

                requested_product_ids = {int(row[0]) for row in product_rows_raw}
                all_product_ids = set(old_quantities) | requested_product_ids

                # Bloquear productos implicados en PostgreSQL.
                product_map: dict[int, dict[str, Any]] = {}
                for product_id in sorted(all_product_ids):
                    product_sql = "SELECT * FROM products WHERE id=?"
                    if db.engine.dialect.name == "postgresql":
                        product_sql += " FOR UPDATE"
                    product = conn.execute(product_sql, (product_id,)).fetchone()
                    if not product:
                        raise ValueError("Uno de los productos ya no existe.")
                    product_map[product_id] = product

                subtotal = Decimal("0")
                product_lines = []
                new_quantities: dict[int, Decimal] = {}

                for product_id_raw, qty_raw, price_raw in product_rows_raw:
                    product_id = int(product_id_raw)
                    qty = parse_decimal(qty_raw)
                    price = parse_decimal(price_raw)

                    if qty <= Decimal("0"):
                        raise ValueError("La cantidad de cada producto debe ser mayor que cero.")
                    if price < Decimal("0"):
                        raise ValueError("El precio de un producto no puede ser negativo.")

                    product = product_map[product_id]
                    line_total = qty * price
                    subtotal += line_total
                    product_lines.append((product, qty, price, line_total))
                    new_quantities[product_id] = new_quantities.get(
                        product_id, Decimal("0")
                    ) + qty

                # Validar y aplicar únicamente la diferencia contra la remisión anterior.
                inventory_adjustments = []
                for product_id in sorted(all_product_ids):
                    product = product_map[product_id]
                    current_stock = Decimal(str(product["stock"] or 0))
                    old_qty = old_quantities.get(product_id, Decimal("0"))
                    new_qty = new_quantities.get(product_id, Decimal("0"))
                    adjustment = old_qty - new_qty
                    new_stock = current_stock + adjustment

                    if new_stock < Decimal("0"):
                        available_for_edit = current_stock + old_qty
                        raise ValueError(
                            f"Stock insuficiente para {product['description']}. "
                            f"Disponible considerando la remisión actual: {available_for_edit:g}."
                        )

                    if adjustment != Decimal("0"):
                        inventory_adjustments.append(
                            (product_id, product, adjustment, new_stock)
                        )

                movement_time = datetime.now().isoformat(timespec="seconds")
                for product_id, product, adjustment, new_stock in inventory_adjustments:
                    conn.execute(
                        """UPDATE products
                           SET stock=?, updated_at=CURRENT_TIMESTAMP
                           WHERE id=?""",
                        (float(new_stock), product_id),
                    )
                    movement_type = (
                        "Ajuste remisión entrada"
                        if adjustment > 0
                        else "Ajuste remisión salida"
                    )
                    conn.execute(
                        """INSERT INTO movements(
                               product_id,movement_type,quantity,reference,notes,created_at,user_id
                           ) VALUES(?,?,?,?,?,?,?)""",
                        (
                            product_id,
                            movement_type,
                            float(abs(adjustment)),
                            current_invoice["number"],
                            "Edición de remisión realizada por Arturo López",
                            movement_time,
                            g.user["id"],
                        ),
                    )

                conn.execute(
                    "DELETE FROM invoice_items WHERE invoice_id=?",
                    (invoice_id,),
                )
                for product, qty, price, line_total in product_lines:
                    conn.execute(
                        """INSERT INTO invoice_items(
                               invoice_id,product_id,code,description,quantity,unit_price,line_total
                           ) VALUES(?,?,?,?,?,?,?)""",
                        (
                            invoice_id,
                            product["id"],
                            product["code"],
                            product["description"],
                            float(qty),
                            float(price),
                            float(line_total),
                        ),
                    )

                # Reconstruir servicios.
                service_lines = []
                for service_id_raw, worker_name, qty_raw, price_raw in service_rows_raw:
                    service_id = int(service_id_raw)
                    worker_name = (worker_name or "").strip()
                    qty = parse_decimal(qty_raw)
                    price = parse_decimal(price_raw)

                    if not worker_name:
                        raise ValueError(
                            "Indica el trabajador responsable de cada servicio."
                        )
                    if qty <= Decimal("0") or price < Decimal("0"):
                        raise ValueError(
                            "Revisa las cantidades y precios de los servicios."
                        )

                    service = conn.execute(
                        "SELECT * FROM services WHERE id=?",
                        (service_id,),
                    ).fetchone()
                    if not service:
                        raise ValueError("Uno de los servicios ya no existe.")

                    line_total = qty * price
                    worker_pct = Decimal(str(service["worker_percentage"] or 70))
                    business_pct = Decimal(str(service["business_percentage"] or 30))
                    worker_amount = (
                        line_total * worker_pct / Decimal("100")
                    ).quantize(Decimal("0.01"))
                    business_amount = line_total - worker_amount
                    subtotal += line_total

                    service_lines.append(
                        (
                            service,
                            worker_name,
                            qty,
                            price,
                            line_total,
                            worker_pct,
                            business_pct,
                            worker_amount,
                            business_amount,
                        )
                    )

                conn.execute(
                    "DELETE FROM invoice_service_items WHERE invoice_id=?",
                    (invoice_id,),
                )
                for (
                    service,
                    worker_name,
                    qty,
                    price,
                    line_total,
                    worker_pct,
                    business_pct,
                    worker_amount,
                    business_amount,
                ) in service_lines:
                    conn.execute(
                        """INSERT INTO invoice_service_items(
                               invoice_id,service_id,service_name,worker_name,quantity,unit_price,line_total,
                               worker_percentage,business_percentage,worker_amount,business_amount
                           ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            invoice_id,
                            service["id"],
                            service["name"],
                            worker_name,
                            float(qty),
                            float(price),
                            float(line_total),
                            float(worker_pct),
                            float(business_pct),
                            float(worker_amount),
                            float(business_amount),
                        ),
                    )

                settings = get_settings(conn)
                tax_rate = Decimal(str(settings["tax_rate"] or 0))
                tax = (subtotal * tax_rate / Decimal("100")).quantize(Decimal("0.01"))
                total = subtotal + tax

                # Cliente: conservar/actualizar o volver a consumidor final.
                customer_id = None
                if customer_name.lower() != "consumidor final" or customer_document:
                    existing = None
                    if customer_document:
                        existing = conn.execute(
                            "SELECT id FROM customers WHERE document=? LIMIT 1",
                            (customer_document,),
                        ).fetchone()

                    if existing:
                        customer_id = existing["id"]
                        conn.execute(
                            """UPDATE customers
                               SET name=?,phone=?,email=?,address=?
                               WHERE id=?""",
                            (
                                customer_name,
                                customer_phone,
                                customer_email,
                                customer_address,
                                customer_id,
                            ),
                        )
                    else:
                        cur_customer = conn.execute(
                            """INSERT INTO customers(document,name,phone,email,address)
                               VALUES(?,?,?,?,?)""",
                            (
                                customer_document,
                                customer_name,
                                customer_phone,
                                customer_email,
                                customer_address,
                            ),
                        )
                        customer_id = cur_customer.lastrowid

                # Cartera: conservar pagos ya realizados y recalcular saldo/estado.
                receivable = conn.execute(
                    "SELECT * FROM accounts_receivable WHERE invoice_id=?",
                    (invoice_id,),
                ).fetchone()
                paid_amount = Decimal("0")
                if receivable:
                    paid_row = conn.execute(
                        """SELECT COALESCE(SUM(amount),0) AS paid
                           FROM accounts_receivable_payments
                           WHERE receivable_id=?""",
                        (receivable["id"],),
                    ).fetchone()
                    paid_amount = Decimal(str((paid_row or {}).get("paid") or 0))

                if payment_method == "Crédito":
                    if paid_amount > total:
                        raise ValueError(
                            "No se puede reducir el total de la remisión por debajo "
                            "de los pagos que ya fueron registrados."
                        )

                    if paid_amount >= total and total > Decimal("0"):
                        credit_status = "PAGADA"
                    elif paid_amount > Decimal("0"):
                        credit_status = "PAGO PARCIAL"
                    else:
                        credit_status = "PENDIENTE"

                    if credit_due_date and paid_amount < total:
                        due_date = datetime.strptime(
                            credit_due_date, "%Y-%m-%d"
                        ).date()
                        if due_date < datetime.now().date():
                            credit_status = "VENCIDA"

                    closed_at = (
                        datetime.now().isoformat(timespec="seconds")
                        if credit_status == "PAGADA"
                        else None
                    )

                    if receivable:
                        conn.execute(
                            """UPDATE accounts_receivable
                               SET original_amount=?,due_date=?,status=?,closed_at=?
                               WHERE id=?""",
                            (
                                float(total),
                                credit_due_date or None,
                                credit_status,
                                closed_at,
                                receivable["id"],
                            ),
                        )
                    else:
                        conn.execute(
                            """INSERT INTO accounts_receivable(
                                   invoice_id,original_amount,due_date,status,created_at,closed_at
                               ) VALUES(?,?,?,?,?,?)""",
                            (
                                invoice_id,
                                float(total),
                                credit_due_date or None,
                                credit_status,
                                datetime.now().isoformat(timespec="seconds"),
                                closed_at,
                            ),
                        )
                elif receivable:
                    if paid_amount > Decimal("0"):
                        raise ValueError(
                            "Esta remisión a crédito ya tiene pagos registrados. "
                            "No puede cambiarse a otra forma de pago desde la edición."
                        )
                    conn.execute(
                        "DELETE FROM accounts_receivable_payments WHERE receivable_id=?",
                        (receivable["id"],),
                    )
                    conn.execute(
                        "DELETE FROM accounts_receivable WHERE id=?",
                        (receivable["id"],),
                    )

                conn.execute(
                    """UPDATE invoices
                       SET customer_id=?,customer_name=?,customer_document=?,customer_phone=?,
                           customer_address=?,vehicle_plate=?,salesperson_id=?,salesperson_name=?,
                           issued_at=?,subtotal=?,tax=?,total=?,payment_method=?,notes=?
                       WHERE id=?""",
                    (
                        customer_id,
                        customer_name,
                        customer_document,
                        customer_phone,
                        customer_address,
                        vehicle_plate,
                        salesperson["id"],
                        salesperson["name"],
                        issued_at,
                        float(subtotal),
                        float(tax),
                        float(total),
                        payment_method,
                        notes,
                        invoice_id,
                    ),
                )

            generate_invoice_pdf(invoice_id)
            flash(
                f"Remisión {invoice['number']} actualizada correctamente.",
                "success",
            )
            return redirect(url_for("invoice_view", invoice_id=invoice_id))

        except (ValueError, KeyError, TypeError) as exc:
            flash(str(exc), "danger")

    with db_conn() as conn:
        invoice = conn.execute(
            "SELECT * FROM invoices WHERE id=?",
            (invoice_id,),
        ).fetchone()
        items = conn.execute(
            "SELECT * FROM invoice_items WHERE invoice_id=? ORDER BY id",
            (invoice_id,),
        ).fetchall()
        service_items = conn.execute(
            "SELECT * FROM invoice_service_items WHERE invoice_id=? ORDER BY id",
            (invoice_id,),
        ).fetchall()
        products = conn.execute(
            """SELECT id,code,description,stock,sale_price,unit
               FROM products
               WHERE active=1
               ORDER BY description"""
        ).fetchall()
        services = conn.execute(
            "SELECT * FROM services WHERE active=1 ORDER BY name"
        ).fetchall()
        salespeople = conn.execute(
            "SELECT id,name FROM salespeople WHERE active=1 ORDER BY name"
        ).fetchall()
        receivable = conn.execute(
            "SELECT * FROM accounts_receivable WHERE invoice_id=?",
            (invoice_id,),
        ).fetchone()

    return render_template(
        "invoice_edit.html",
        invoice=invoice,
        items=items,
        service_items=service_items,
        products=products,
        services=services,
        salespeople=salespeople,
        receivable=receivable,
    )


@app.route("/facturas/<int:invoice_id>/pdf")
@login_required
def invoice_pdf(invoice_id: int):
    pdf_path = generate_invoice_pdf(invoice_id)
    return send_file(pdf_path, as_attachment=True, download_name=pdf_path.name)


def generate_invoice_pdf(invoice_id: int) -> Path:
    with db_conn() as conn:
        invoice = conn.execute("""SELECT i.*, u.full_name AS user_name FROM invoices i
            LEFT JOIN users u ON u.id=i.user_id WHERE i.id=?""", (invoice_id,)).fetchone()
        items = conn.execute("SELECT * FROM invoice_items WHERE invoice_id=?", (invoice_id,)).fetchall()
        service_items = conn.execute("SELECT * FROM invoice_service_items WHERE invoice_id=?", (invoice_id,)).fetchall()
        settings = get_settings(conn)
    if not invoice:
        raise ValueError("Remisión no encontrada")

    path = INVOICE_DIR / f"Remision_{invoice['number'].replace('-', '_')}.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=letter, rightMargin=16*mm, leftMargin=16*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=8.5, leading=11))
    story = []

    header = Table([
        [Paragraph(f"<b>{settings['business_name']}</b><br/>{settings['nit']}<br/>{settings['address']}<br/>{settings['phone']} - {settings['email']}", styles["Small"]),
         Paragraph(f"<b>REMISIÓN</b><br/><font size=13>{invoice['number']}</font><br/>Fecha: {invoice['issued_at'][:10]}<br/>Estado: {invoice['status']}", styles["Small"])]
    ], colWidths=[110*mm, 70*mm])
    header.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.8, colors.HexColor("#1F4E78")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BACKGROUND", (1,0), (1,0), colors.HexColor("#D9EAF7")),
        ("LEFTPADDING", (0,0), (-1,-1), 8), ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 8), ("BOTTOMPADDING", (0,0), (-1,-1), 8),
    ]))
    story.extend([header, Spacer(1, 6*mm)])

    customer = Table([
        ["Cliente", invoice["customer_name"], "Documento", invoice["customer_document"] or "-"],
        ["Teléfono", invoice["customer_phone"] or "-", "Dirección", invoice["customer_address"] or "-"],
        ["Placa", invoice["vehicle_plate"] or "-", "Forma de pago", invoice["payment_method"]],
        ["Vendedor", invoice["salesperson_name"] or "-", "Usuario", invoice["user_name"] or "-"],
        ["Observaciones", invoice["notes"] or "-", "", ""],
    ], colWidths=[25*mm, 65*mm, 28*mm, 62*mm])
    customer.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.4, colors.grey),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#EEF3F8")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#EEF3F8")),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"), ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8), ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("PADDING", (0,0), (-1,-1), 5),
    ]))
    story.extend([customer, Spacer(1, 6*mm)])

    data = [["Tipo", "Descripción", "Cant.", "Valor unitario", "Total"]]
    for item in items:
        data.append([
            "Producto", Paragraph(item["description"], styles["Small"]), f"{item['quantity']:g}",
            money(item["unit_price"]), money(item["line_total"])
        ])
    for item in service_items:
        data.append([
            "Servicio", Paragraph(item["service_name"], styles["Small"]), f"{item['quantity']:g}",
            money(item["unit_price"]), money(item["line_total"])
        ])
    data.extend([
        ["", "", "", "Subtotal", money(invoice["subtotal"])],
        ["", "", "", "Impuesto", money(invoice["tax"])],
        ["", "", "", "TOTAL", money(invoice["total"])],
    ])
    table = Table(data, colWidths=[25*mm, 78*mm, 16*mm, 30*mm, 31*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ("GRID", (0,0), (-1,-4), 0.35, colors.grey),
        ("LINEABOVE", (3,-3), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (3,-1), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (3,-1), (-1,-1), colors.HexColor("#D9EAF7")),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.extend([table, Spacer(1, 9*mm), Paragraph("Documento de remisión. Gracias por su compra.", styles["Normal"])])
    doc.build(story)
    return path


@app.route("/configuracion", methods=["GET", "POST"])
@role_required("administrador")
def settings_view():
    if request.method == "POST":
        try:
            tax_rate = float(parse_decimal(request.form.get("tax_rate")))
            with db_conn() as conn:
                conn.execute(
                    """UPDATE settings SET business_name=?,nit=?,address=?,phone=?,email=?,invoice_prefix=?,tax_rate=? WHERE id=1""",
                    (request.form["business_name"].strip(), request.form.get("nit", "").strip(),
                     request.form.get("address", "").strip(), request.form.get("phone", "").strip(),
                     request.form.get("email", "").strip(), request.form.get("invoice_prefix", "REM").strip().upper(), tax_rate),
                )
            flash("Configuración guardada.", "success")
            return redirect(url_for("settings_view"))
        except ValueError as exc:
            flash(str(exc), "danger")
    settings = get_settings()
    return render_template("settings.html", settings=settings, db_path=DB_PATH, db_backend=db.engine.dialect.name)


@app.route("/respaldo")
@role_required("administrador")
def backup():
    """Exporta un respaldo JSON portable tanto para SQLite como PostgreSQL."""
    import json
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = BACKUP_DIR / f"inventario_respaldo_{stamp}.json"
    table_names = [
        "users", "settings", "products", "customers", "salespeople", "services",
        "invoices", "invoice_items", "invoice_service_items", "movements",
        "accounts_receivable", "accounts_receivable_payments"
    ]
    payload = {"created_at": datetime.now().isoformat(), "database": db.engine.dialect.name, "tables": {}}
    with db_conn() as conn:
        for table_name in table_names:
            payload["tables"][table_name] = conn.execute(f"SELECT * FROM {table_name}").fetchall()
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return send_file(path, as_attachment=True, download_name=path.name)


def open_browser():
    time.sleep(1.2)
    webbrowser.open("http://127.0.0.1:5000")


# En Render/Gunicorn el módulo se importa, por eso inicializamos también en import.
# La operación es idempotente: no borra productos, usuarios, vendedores ni servicios existentes.
init_db()

if __name__ == "__main__":
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host="127.0.0.1", port=int(os.getenv("PORT", "5000")), debug=False, use_reloader=False)
