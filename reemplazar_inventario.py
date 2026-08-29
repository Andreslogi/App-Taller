from __future__ import annotations

import argparse
import json
import os
from collections import Counter, OrderedDict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "Inventario"
EXPECTED_HEADERS = [
    "Categoría",
    "Producto",
    "Referencia / Descripción",
    "Cantidad",
    "Precio unitario",
    "Valor total",
]


def clean_text(value) -> str:
    """Conserva el texto del Excel, eliminando solo espacios exteriores accidentales."""
    if value is None:
        return ""
    return str(value).strip()


def as_decimal(value, *, field: str, allow_blank: bool = False) -> Decimal:
    if value is None or clean_text(value) == "":
        if allow_blank:
            return Decimal("0")
        raise ValueError(f"El campo '{field}' está vacío.")

    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    raw = clean_text(value).replace("$", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")

    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Valor inválido en '{field}': {value!r}") from exc


def normalize_price(value) -> Decimal:
    # Precio vacío = pendiente. La BD exige un número, por eso se guarda 0.
    return as_decimal(value, field="Precio unitario", allow_blank=True)


def read_inventory(excel_path: Path) -> list[dict]:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"No existe la hoja '{SHEET_NAME}'. Hojas encontradas: {', '.join(wb.sheetnames)}"
        )

    ws = wb[SHEET_NAME]

    headers = [clean_text(ws.cell(3, col).value) for col in range(1, 7)]
    if headers != EXPECTED_HEADERS:
        raise ValueError(
            "Los encabezados del Excel no coinciden con lo esperado.\n"
            f"Esperados: {EXPECTED_HEADERS}\n"
            f"Encontrados: {headers}"
        )

    # Clave exacta de producto: categoría + producto + referencia.
    # Esto permite unir Rodillo 30213 pero mantiene separados, por ejemplo,
    # Juego completo y Medio juego de una misma banda.
    grouped: OrderedDict[tuple[str, str, str], dict] = OrderedDict()

    for row_number in range(4, ws.max_row + 1):
        category = clean_text(ws.cell(row_number, 1).value)
        product = clean_text(ws.cell(row_number, 2).value)
        reference = clean_text(ws.cell(row_number, 3).value)
        qty_raw = ws.cell(row_number, 4).value
        price_raw = ws.cell(row_number, 5).value

        # Ignorar filas vacías y la fila TOTAL GENERAL.
        if category.upper() == "TOTAL GENERAL":
            break
        if not any([category, product, reference, qty_raw, price_raw]):
            continue

        if not product:
            raise ValueError(f"Fila {row_number}: el campo Producto está vacío.")
        if not category:
            raise ValueError(f"Fila {row_number}: la Categoría está vacía para '{product}'.")
        if not reference:
            # Si falta referencia, usamos el producto como referencia sin cambiar el nombre del producto.
            reference = product

        quantity = as_decimal(qty_raw, field=f"Cantidad (fila {row_number})", allow_blank=True)
        if quantity < 0:
            raise ValueError(f"Fila {row_number}: la cantidad no puede ser negativa ({product}).")

        price = normalize_price(price_raw)
        if price < 0:
            raise ValueError(f"Fila {row_number}: el precio no puede ser negativo ({product}).")

        key = (category, product, reference)

        if key in grouped:
            current = grouped[key]
            # Un duplicado exacto solo se puede unir si el precio coincide.
            if current["sale_price"] != price:
                raise ValueError(
                    "Duplicado con precios diferentes. Se detuvo la carga para evitar pérdidas:\n"
                    f"Producto: {product}\nReferencia: {reference}\n"
                    f"Precio 1: {current['sale_price']}\nPrecio 2: {price}\nFila: {row_number}"
                )
            current["stock"] += quantity
            current["source_rows"].append(row_number)
        else:
            grouped[key] = {
                "category": category,
                "product": product,
                "reference": reference,
                "stock": quantity,
                "sale_price": price,
                "source_rows": [row_number],
            }

    items = list(grouped.values())
    if not items:
        raise ValueError("No se encontraron productos válidos en el Excel.")

    # La tabla products exige code único.
    # Usamos la referencia tal cual cuando es única. Si la misma referencia se usa
    # para varios productos, agregamos el nombre del producto para evitar conflicto.
    reference_counts = Counter(item["reference"] for item in items)
    used_codes: set[str] = set()

    for item in items:
        reference = item["reference"]
        product = item["product"]

        if reference_counts[reference] == 1:
            code = reference
        else:
            code = f"{reference} | {product}"

        # Protección adicional por si el Excel contiene casos excepcionalmente iguales.
        base_code = code[:100]
        code = base_code
        suffix = 2
        while code.casefold() in used_codes:
            suffix_text = f" #{suffix}"
            code = (base_code[: 100 - len(suffix_text)] + suffix_text)
            suffix += 1

        used_codes.add(code.casefold())
        item["code"] = code

    return items


def database_label() -> str:
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        return "SQLite local (DATABASE_URL no definido)"
    # No mostrar credenciales.
    if "@" in url:
        return url.split("@", 1)[1]
    return url


def create_backup(conn, backup_dir: Path, db) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = backup_dir / f"antes_reemplazo_inventario_{stamp}.json"

    tables = [
        "users",
        "settings",
        "products",
        "customers",
        "salespeople",
        "services",
        "invoices",
        "invoice_items",
        "invoice_service_items",
        "movements",
    ]

    payload = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "database_backend": db.engine.dialect.name,
        "tables": {},
    }

    for table in tables:
        payload["tables"][table] = conn.execute(f"SELECT * FROM {table}").fetchall()

    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return path


def preview(items: list[dict]) -> None:
    total_stock = sum((item["stock"] for item in items), Decimal("0"))
    missing_prices = [item for item in items if item["sale_price"] == 0]
    low = [item for item in items if item["stock"] == 1]
    out = [item for item in items if item["stock"] <= 0]
    merged = [item for item in items if len(item["source_rows"]) > 1]

    print("\n========== VALIDACIÓN DEL EXCEL ==========")
    print(f"Productos finales a cargar : {len(items)}")
    print(f"Unidades totales           : {total_stock}")
    print(f"Productos sin precio       : {len(missing_prices)}")
    print(f"Productos con stock = 1    : {len(low)}")
    print(f"Productos agotados         : {len(out)}")
    print(f"Duplicados unidos          : {len(merged)}")

    if merged:
        print("\nDuplicados que se unirán:")
        for item in merged:
            print(
                f"  - {item['product']} | {item['reference']} | "
                f"stock final={item['stock']} | filas={item['source_rows']}"
            )

    if missing_prices:
        print("\nProductos sin precio (se cargarán con precio 0):")
        for item in missing_prices:
            print(f"  - {item['product']} | stock={item['stock']}")

    print("==========================================\n")


def replace_inventory(items: list[dict], backup_dir: Path, reset_invoice_number: bool) -> Path:
    from database import db, db_conn

    with db_conn() as conn:
        backup_path = create_backup(conn, backup_dir, db)

        # Limpiar únicamente la información de inventario/remisiones de prueba.
        # El catálogo de servicios NO se borra.
        conn.execute("DELETE FROM invoice_service_items")
        conn.execute("DELETE FROM invoice_items")
        conn.execute("DELETE FROM movements")
        conn.execute("DELETE FROM invoices")
        conn.execute("DELETE FROM products")

        if reset_invoice_number:
            conn.execute("UPDATE settings SET next_invoice=1 WHERE id=1")

        for item in items:
            # Precio vacío => sale_price=0 y cost=0.
            # stock mínimo siempre 1, como pidió el usuario.
            conn.execute(
                """
                INSERT INTO products(
                    code,
                    description,
                    category,
                    unit,
                    stock,
                    min_stock,
                    cost,
                    sale_price,
                    notes,
                    active
                )
                VALUES(?,?,?,?,?,?,?,?,?,1)
                """,
                (
                    item["code"],
                    item["product"],
                    item["category"],
                    "Unidad",
                    float(item["stock"]),
                    1.0,
                    0.0,
                    float(item["sale_price"]),
                    item["reference"],
                ),
            )

        check = conn.execute(
            """
            SELECT
                COUNT(*) AS products,
                COALESCE(SUM(stock),0) AS stock,
                SUM(CASE WHEN stock <= 0 THEN 1 ELSE 0 END) AS agotados,
                SUM(CASE WHEN stock > 0 AND stock <= min_stock THEN 1 ELSE 0 END) AS bajos,
                SUM(CASE WHEN sale_price <= 0 THEN 1 ELSE 0 END) AS sin_precio
            FROM products
            WHERE active=1
            """
        ).fetchone()

        if int(check["products"] or 0) != len(items):
            raise RuntimeError(
                f"Validación final falló: se esperaban {len(items)} productos y quedaron {check['products']}."
            )

        print("\n========== RESULTADO EN BASE DE DATOS ==========")
        print(f"Productos cargados : {check['products']}")
        print(f"Stock total        : {check['stock']}")
        print(f"Bajo inventario    : {check['bajos']}")
        print(f"Agotados           : {check['agotados']}")
        print(f"Sin precio         : {check['sin_precio']}")
        print("Servicios          : CONSERVADOS")
        print("=================================================\n")

        return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reemplaza completamente el inventario desde Inventario_Repuestos.xlsx sin borrar el catálogo de servicios."
    )
    parser.add_argument("excel", help="Ruta al archivo Inventario_Repuestos.xlsx")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Realiza el reemplazo. Sin esta opción solo valida y muestra un resumen.",
    )
    parser.add_argument(
        "--confirm",
        default="",
        help="Para aplicar debe ser exactamente REEMPLAZAR_INVENTARIO",
    )
    parser.add_argument(
        "--keep-invoice-number",
        action="store_true",
        help="No reinicia el consecutivo de remisiones a 1.",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    items = read_inventory(excel_path)
    preview(items)

    print(f"Base objetivo: {database_label()}")

    if not args.apply:
        print("MODO VALIDACIÓN: no se modificó ninguna base de datos.")
        print("Si todo está correcto, vuelve a ejecutar con --apply --confirm REEMPLAZAR_INVENTARIO")
        return

    if args.confirm != "REEMPLAZAR_INVENTARIO":
        raise SystemExit(
            "SEGURIDAD: para modificar la base agrega: --confirm REEMPLAZAR_INVENTARIO"
        )

    from app import app

    with app.app_context():
        backup_dir = Path(__file__).resolve().parent / "backups"
        backup_path = replace_inventory(
            items,
            backup_dir=backup_dir,
            reset_invoice_number=not args.keep_invoice_number,
        )

    print("INVENTARIO REEMPLAZADO CORRECTAMENTE.")
    print(f"Respaldo creado en: {backup_path}")


if __name__ == "__main__":
    # app.py ya inicializa Flask/SQLAlchemy al importarse. Este main solo ejecuta la carga.
    # La transacción de db_conn hace rollback automático si ocurre cualquier error.
    main()
